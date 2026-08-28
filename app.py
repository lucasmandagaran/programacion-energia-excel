from __future__ import annotations

import hashlib
import io
import json
import os
import re
import unicodedata
from datetime import date, datetime, timezone
from typing import Any
from zoneinfo import ZoneInfo

import pandas as pd
import pdfplumber
import requests
import streamlit as st
from openpyxl.styles import Font, PatternFill


APP_TITLE = "Programacion Energia"
LOCAL_TZ = ZoneInfo("America/Argentina/Buenos_Aires")
ADVANCE_EXPORT_COLUMNS = {
    "estado",
    "modificaciones programa",
    "comentario",
    "fecha inicio tarea",
    "fecha finalizacion tarea",
    "fecha avance",
    "hora avance",
    "fecha modificacion",
    "hora modificacion",
    "informado por",
    "empresa informante",
}
AREAS = ["GENERACION", "DISTRIBUCION"]
COMPANIES_BY_AREA = {
    "GENERACION": ["", "MANPETROL", "SAN&FRAN", "OTRA"],
    "DISTRIBUCION": ["", "MANPETROL", "ELECTRO PATAGONIA", "OTRA"],
}
SECTORS_BY_AREA = {
    "GENERACION": ["", "Electricidad", "Mecanica", "Instrumentacion", "Otros"],
    "DISTRIBUCION": [
        "",
        "Estaciones Transformadoras",
        "Telemando",
        "Trabajos con tension",
        "Trabajos sin tension",
        "Guardia 24hs",
        "Otros",
    ],
}
COMPANIES = COMPANIES_BY_AREA["GENERACION"]
SECTORS = SECTORS_BY_AREA["GENERACION"]
GENERATION_SECTORS = {"electricidad", "mecanica", "instrumentacion"}
DISTRIBUTION_SECTORS = {
    "estaciones transformadoras",
    "telemando",
    "trabajos con tension",
    "trabajos sin tension",
    "guardia 24hs",
}
DISTRIBUTION_CREW_SCOPES = {
    # Cuadrillas de Distribucion. Estas reglas tienen prioridad sobre la
    # carpeta/sector del Excel porque un mismo sector puede tener varias empresas.
    "600": ("ELECTRO PATAGONIA", "Estaciones Transformadoras"),
    "612": ("ELECTRO PATAGONIA", "Estaciones Transformadoras"),
    "614": ("MANPETROL", "Estaciones Transformadoras"),
    "618": ("ELECTRO PATAGONIA", "Telemando"),
    "652": ("MANPETROL", "Trabajos con tension"),
    "653": ("MANPETROL", "Trabajos con tension"),
    "654": ("MANPETROL", "Trabajos con tension"),
    "655": ("MANPETROL", "Trabajos con tension"),
    "656": ("MANPETROL", "Trabajos con tension"),
    "657": ("MANPETROL", "Trabajos con tension"),
    "658": ("MANPETROL", "Trabajos con tension"),
    "659": ("MANPETROL", "Trabajos con tension"),
    "601": ("ELECTRO PATAGONIA", "Trabajos sin tension"),
    "603": ("ELECTRO PATAGONIA", "Trabajos sin tension"),
    "615": ("ELECTRO PATAGONIA", "Trabajos sin tension"),
    "616": ("ELECTRO PATAGONIA", "Trabajos sin tension"),
    "617": ("ELECTRO PATAGONIA", "Trabajos sin tension"),
    "619": ("ELECTRO PATAGONIA", "Trabajos sin tension"),
    "621": ("ELECTRO PATAGONIA", "Trabajos sin tension"),
    "627": ("ELECTRO PATAGONIA", "Trabajos sin tension"),
    "G24": ("MANPETROL", "Guardia 24hs"),
    "502A": ("MANPETROL", "Guardia 24hs"),
}
STATE_ACTIONS = ["EN CURSO", "EN ESPERA", "COMPLETADO", "REPLANIFICAR", "SIN AVANCE"]
REASON_ACTIONS = {"EN ESPERA", "REPLANIFICAR"}
HIDE_AFTER_SAVE_ACTIONS = {"COMPLETADO", "REPLANIFICAR"}
CREW_OPERATIONAL_STATES = ["Operativa", "No operativa"]
DATE_MODE_TODAY = "Fecha de hoy"
DATE_MODE_PROGRAM = "Fecha segun programa"
DATE_MODE_MANUAL = "Fecha manual"
TASK_DATE_MODES = [DATE_MODE_TODAY, DATE_MODE_PROGRAM, DATE_MODE_MANUAL]
SEARCH_FIELD_OPTIONS = {
    "Titulo tarea": ["tarea"],
    "OT": ["nro_ot"],
    "Ubicacion tecnica": ["ubicacion_tecnica"],
    "KKS/TAG": ["kks_tag"],
    "Cuadrilla": ["cuadrilla"],
}
STATE_ROW_STYLES = {
    "COMPLETADO": "background-color: rgba(36, 161, 72, 0.30); color: #f2fff5; font-weight: 700;",
    "EN CURSO": "background-color: rgba(31, 111, 235, 0.28); color: #f2f7ff; font-weight: 700;",
    "EN ESPERA": "background-color: rgba(214, 158, 46, 0.30); color: #fff8e5; font-weight: 700;",
    "REPLANIFICAR": "background-color: rgba(248, 81, 73, 0.30); color: #fff0ef; font-weight: 700;",
    "COMENTARIO": "background-color: rgba(137, 87, 229, 0.22); color: #f6efff; font-weight: 700;",
    "SIN AVANCE": "background-color: rgba(139, 148, 158, 0.14); color: #f0f3f6;",
}
COMMENT_DISPLAY_LIMIT = 95
REASONS = [
    "",
    "Pedido Sup PAE",
    "Por factor climatico",
    "Por falta de equipo/recursos/materiales",
    "Cuadrilla no operativa",
    "Otros",
]
OT_COLUMNS = ["ot", "ots", "orden", "ordenes", "orden de trabajo", "nro ot", "nro de ot", "nro. de ot", "numero ot", "numero de ot", "nrodeot", "nroot"]
TITLE_COLUMNS = ["title", "titulo", "titulo tarea", "titulo de tarea", "trabajo", "tarea", "descripcion", "description", "texto breve", "textobreve", "nombre", "nombre tarea"]
COMPANY_COLUMNS = ["empresa", "contratista", "compania", "cia"]
SECTOR_COLUMNS = ["sector", "especialidad", "disciplina", "puesto de trabajo"]
CREW_COLUMNS = ["cuadrilla", "cuadrillas", "cuadrilla gen", "cuadrillagen", "crew", "recurso", "recursos"]
START_DATE_COLUMNS = [
    "fecha inicio",
    "fecha de inicio",
    "fecha inic",
    "fecha ini",
    "fecha inicio extrema",
    "fecha de inicio extrema",
    "inicio",
    "start",
    "start date",
    "fecha programada",
    "fecha de inicio programada",
    "fechaprogramada",
]
END_DATE_COLUMNS = [
    "fecha fin",
    "fecha de fin",
    "fecha fin extrema",
    "fecha de fin extrema",
    "fecha finalizacion",
    "fecha de finalizacion",
    "fecha vencimiento",
    "fecha de vencimiento",
    "fechas vencimiento",
    "fechas de vencimiento",
    "fechas de ve",
    "vencimiento",
    "fin",
    "end",
    "end date",
    "due",
    "due date",
    "fecha cierre",
    "fecha de cierre",
    "cierre",
]
STATUS_COLUMNS = ["estado", "status", "estado actual", "status de usuario", "avance"]
LOCATION_COLUMNS = ["ubicacion tecnica", "ubicacion", "ubic tecnica", "ubic. tecnica", "ubictecnica", "objeto ubicacion", "objetoubicacion"]
KKS_COLUMNS = ["kks tag", "kks-tag", "kks/tag", "kks", "tag", "kkstag", "kks tag ubicacion", "kkstagubicacion"]
CREW_STATUS_CUADRILLA_COLUMNS = ["cuadrilla"]
CREW_STATUS_MOVIL_COLUMNS = ["interno movil", "movil", "interno", "int"]
CREW_STATUS_TETRA_COLUMNS = ["tetra", "contacto"]
CREW_STATUS_ENCARGADO_COLUMNS = ["encargado"]
CREW_STATUS_OPERARIO_COLUMNS = ["operario"]
CREW_STATUS_INTEGRANTES_COLUMNS = ["conformacion de equipo", "conformacion equipo", "equipo", "integrantes"]
CREW_STATUS_DISPONIBILIDAD_COLUMNS = ["disponibilidad"]
CREW_STATUS_NA_VALUES = {"na", "n a"}


def option_label(value: str, empty_label: str) -> str:
    return value or empty_label


def scope_value(value: Any) -> str:
    text = str(value or "").strip()
    if normalize(text) in {"todas", "todos", "all"}:
        return ""
    return text


def canonical_area(value: Any) -> str:
    norm = compact_key(value)
    if "distrib" in norm or norm in {"td", "tyd", "tandd"}:
        return "DISTRIBUCION"
    return "GENERACION"


def current_area() -> str:
    profile = st.session_state.get("profile", {})
    return canonical_area(profile.get("area") or "GENERACION")


def program_area(program: dict[str, Any]) -> str:
    return canonical_area(program.get("area") or "GENERACION")


def task_area(task: dict[str, Any]) -> str:
    return canonical_area(task.get("area") or "GENERACION")


def secret(name: str, default: str = "") -> str:
    try:
        return str(st.secrets.get(name, "") or os.getenv(name, "") or default)
    except Exception:
        return str(os.getenv(name, "") or default)


def required_secret(name: str) -> str:
    value = secret(name)
    if not value:
        st.error(f"Falta configurar {name} en Streamlit Secrets.")
        st.stop()
    return value


SUPABASE_URL = secret("SUPABASE_URL").rstrip("/")
if SUPABASE_URL.endswith("/rest/v1"):
    SUPABASE_URL = SUPABASE_URL[: -len("/rest/v1")]
SUPABASE_KEY = secret("SUPABASE_KEY")
ACCESS_PASSWORD = required_secret("ACCESS_PASSWORD")
ADMIN_PASSWORD = required_secret("ADMIN_PASSWORD")


st.set_page_config(page_title=APP_TITLE, page_icon="PE", layout="wide")


def hide_streamlit_chrome() -> None:
    st.markdown(
        """
        <style>
        :root {
            --pe-gap: 0.45rem;
        }
        #MainMenu,
        footer,
        [data-testid="stToolbar"],
        [data-testid="stDecoration"],
        [data-testid="stHeaderActionElements"],
        [data-testid="stStatusWidget"],
        [data-testid="stLogo"],
        [data-testid="stAppDeployButton"],
        [data-testid="stAppCreatorAvatar"],
        [data-testid="stAppCreatorBadge"],
        [data-testid="stViewerBadge"],
        .stDeployButton,
        .viewerBadge_container__1QSob,
        [class*="viewerBadge"],
        [class*="ViewerBadge"],
        [class*="hostedWithStreamlit"],
        [class*="HostedWithStreamlit"],
        [class*="creatorAvatar"],
        [class*="CreatorAvatar"],
        a[href*="streamlit.io"],
        a[href*="github.com/lucasmandagaran/programacion-energia-excel"] {
            display: none !important;
            visibility: hidden !important;
        }
        header {
            height: 0 !important;
            min-height: 0 !important;
            background: transparent !important;
        }
        .block-container {
            padding-top: 1.05rem !important;
            padding-bottom: 1rem !important;
            max-width: 100% !important;
        }
        h1 {
            font-size: clamp(1.75rem, 3.4vw, 2.55rem) !important;
            line-height: 1.05 !important;
            margin-bottom: 0.2rem !important;
        }
        h2, h3 {
            margin-top: 0.45rem !important;
            margin-bottom: 0.35rem !important;
        }
        [data-testid="stVerticalBlock"] {
            gap: var(--pe-gap) !important;
        }
        [data-testid="stHorizontalBlock"] {
            gap: 0.55rem !important;
        }
        [data-testid="stExpander"] {
            margin-bottom: 0.35rem !important;
        }
        [data-testid="stDataFrame"] {
            margin-top: 0.15rem !important;
        }
        [data-testid="stDataFrame"] div[role="row"]:hover,
        [data-testid="stDataFrame"] div[role="gridcell"]:hover {
            filter: brightness(1.05) !important;
        }
        .stButton > button,
        .stDownloadButton > button {
            min-height: 2.25rem !important;
            padding-top: 0.35rem !important;
            padding-bottom: 0.35rem !important;
        }
        .stTextInput input,
        .stSelectbox div[data-baseweb="select"],
        .stMultiSelect div[data-baseweb="select"],
        .stDateInput input {
            min-height: 2.25rem !important;
        }
        .pe-header {
            margin: 0 0 0.25rem 0;
        }
        .pe-header h1 {
            margin: 0 !important;
        }
        .pe-subtitle {
            margin: 0.05rem 0 0.35rem 0;
            color: rgba(250, 250, 250, 0.68);
            font-size: 0.88rem;
        }
        @media (max-width: 760px) {
            .block-container {
                padding-left: 0.55rem !important;
                padding-right: 0.55rem !important;
                padding-top: 0.55rem !important;
            }
            h1 {
                font-size: 1.65rem !important;
            }
            .stCaptionContainer,
            [data-testid="stCaptionContainer"] {
                font-size: 0.78rem !important;
            }
            [data-testid="column"] {
                width: 100% !important;
                flex: 1 1 100% !important;
            }
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


hide_streamlit_chrome()


def app_header(subtitle: str = "") -> None:
    subtitle_html = f'<div class="pe-subtitle">{subtitle}</div>' if subtitle else ""
    st.markdown(
        f"""
        <div class="pe-header">
            <h1>{APP_TITLE}</h1>
            {subtitle_html}
        </div>
        """,
        unsafe_allow_html=True,
    )


def normalize(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = "".join(
        char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn"
    )
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def compact_key(value: Any) -> str:
    return normalize(value).replace(" ", "")


def normalize_crew(value: Any) -> str:
    # Antes del "/" suele venir un sufijo de turno (ej.: "601/12h", "627/24hs")
    # que no forma parte del numero de cuadrilla.
    base = str(value or "").split("/")[0]
    text = normalize(base).upper().replace(" ", "")
    text = re.sub(r"(MP|SF|SANFRAN|MANPETROL)$", "", text)
    return text


def canonical_company(value: Any) -> str:
    norm = compact_key(value)
    if not norm:
        return ""
    if "manpetrol" in norm or norm in {"mp", "manpet"}:
        return "MANPETROL"
    if "sanfran" in norm or "sanyfran" in norm or norm in {"sf", "sanf"}:
        return "SAN&FRAN"
    if "electropatagonia" in norm or norm in {"ep", "elepatagonia"}:
        return "ELECTRO PATAGONIA"
    if norm in {"otra", "otro", "otros", "otras"}:
        return "OTRA"
    return str(value or "").strip()


def canonical_sector(value: Any) -> str:
    norm = compact_key(value)
    if not norm:
        return ""
    company, distribution_sector = distribution_company_sector(value)
    if distribution_sector:
        return distribution_sector
    if "elect" in norm or norm in {"elec", "pdgelec", "pdegelec", "pdgelect", "pdgelectrico"}:
        return "Electricidad"
    if "inst" in norm or norm in {"pdginst", "pdeginst", "pdginstrumentacion"}:
        return "Instrumentacion"
    if "mec" in norm or "mecan" in norm or norm in {"pdgmec", "pdegmeca", "pdgmecanica"}:
        return "Mecanica"
    if norm in {"otro", "otros", "otras"}:
        return "Otros"
    return str(value or "").strip()


def distribution_company_sector(value: Any) -> tuple[str, str]:
    norm = compact_key(value)
    if not norm:
        return "", ""
    if "tareasrelevantes" in norm:
        return "", ""
    if "pdedeett" in norm or "estacionestransformadoras" in norm:
        return "MANPETROL", "Estaciones Transformadoras"
    if "pdedstel" in norm or "telemando" in norm:
        return "ELECTRO PATAGONIA", "Telemando"
    if "pdedetct" in norm or norm.startswith("tct") or "trabajoscontension" in norm:
        return "MANPETROL", "Trabajos con tension"
    if "pdedstop" in norm or norm.startswith("tst") or "trabajossintension" in norm:
        return "ELECTRO PATAGONIA", "Trabajos sin tension"
    if "guardia24" in norm:
        return "MANPETROL", "Guardia 24hs"
    return "", ""


def distribution_company_sector_from_crew(cuadrilla: Any) -> tuple[str, str]:
    return DISTRIBUTION_CREW_SCOPES.get(normalize_crew(cuadrilla), ("", ""))


def distribution_sector_company(sector: Any) -> str:
    company, _ = distribution_company_sector(sector)
    return company


def crew_context_lookup(tasks: list[dict[str, Any]]) -> dict[str, tuple[str, str]]:
    lookup: dict[str, tuple[str, str]] = {}
    for task in tasks:
        crew_key = normalize_crew(task.get("cuadrilla"))
        if not crew_key or crew_key in lookup:
            continue
        company = effective_company(task)
        sector = effective_sector(task)
        if company or sector:
            lookup[crew_key] = (company, sector)
    return lookup


def infer_company_sector(cuadrilla: Any, area: Any = "GENERACION", sector_value: Any = "") -> tuple[str, str]:
    if canonical_area(area) == "DISTRIBUCION":
        crew_company, crew_sector = distribution_company_sector_from_crew(cuadrilla)
        if crew_company or crew_sector:
            return crew_company, crew_sector
        return distribution_company_sector(sector_value)
    crew = normalize_crew(cuadrilla)
    if crew in {"555", "555A", "A555"}:
        return "MANPETROL", "Electricidad"
    if crew in {"556A", "556B", "556C"}:
        return "MANPETROL", "Instrumentacion"
    if crew in {"720", "721", "722", "723", "724"}:
        return "SAN&FRAN", "Mecanica"
    return "", "Otros"


def effective_company(task: dict[str, Any]) -> str:
    area = task_area(task)
    if area == "DISTRIBUCION":
        crew_status_company = canonical_company(task.get("_crew_status_company"))
        if crew_status_company:
            return crew_status_company
    value = canonical_company(task.get("empresa"))
    inferred, _ = infer_company_sector(task.get("cuadrilla"), area, task.get("sector"))
    crew_company = canonical_company(task.get("cuadrilla"))
    return value or inferred or crew_company


def effective_sector(task: dict[str, Any]) -> str:
    area = task_area(task)
    if area == "DISTRIBUCION":
        crew_status_sector = canonical_sector(task.get("_crew_status_sector"))
        if crew_status_sector:
            return crew_status_sector
    value = canonical_sector(task.get("sector"))
    _, inferred = infer_company_sector(task.get("cuadrilla"), area, task.get("sector"))
    return value or inferred


def is_other_company_scope(task: dict[str, Any]) -> bool:
    area = task_area(task)
    inferred_company, _ = infer_company_sector(task.get("cuadrilla"), area, task.get("sector"))
    return not inferred_company


def is_other_sector_scope(task: dict[str, Any]) -> bool:
    area = task_area(task)
    _, inferred_sector = infer_company_sector(task.get("cuadrilla"), area, task.get("sector"))
    if inferred_sector == "Otros":
        return True
    known = GENERATION_SECTORS if area == "GENERACION" else DISTRIBUTION_SECTORS
    return normalize(effective_sector(task)) not in known


def supabase_headers(prefer: str | None = None) -> dict[str, str]:
    if not SUPABASE_URL or not SUPABASE_KEY:
        st.error("Falta configurar SUPABASE_URL y SUPABASE_KEY en Streamlit Secrets.")
        st.stop()
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    return headers


def api_url(table: str) -> str:
    return f"{SUPABASE_URL}/rest/v1/{table}"


def stop_supabase_error(response: requests.Response, action: str) -> None:
    message = response.text.strip()
    try:
        payload = response.json()
        message = json.dumps(payload, ensure_ascii=False, indent=2)
    except Exception:
        pass
    st.error(f"Error de Supabase al {action}. Codigo HTTP: {response.status_code}")
    if response.status_code in {401, 403}:
        st.warning("Revisar SUPABASE_KEY y los permisos/grants del SQL.")
    elif response.status_code == 404:
        st.warning("Revisar que existan las tablas public.programs, public.tasks y public.advances.")
    st.code(message[:2000] or "Sin detalle devuelto por Supabase.")
    st.stop()


def sb_get(table: str, params: dict[str, str] | None = None) -> list[dict[str, Any]]:
    try:
        response = requests.get(api_url(table), headers=supabase_headers(), params=params or {}, timeout=30)
    except requests.RequestException as exc:
        st.error("No se pudo conectar con Supabase.")
        st.code(str(exc))
        st.stop()
    if not response.ok:
        stop_supabase_error(response, f"leer {table}")
    return response.json()


def sb_insert(table: str, rows: list[dict[str, Any]] | dict[str, Any]) -> list[dict[str, Any]]:
    try:
        response = requests.post(
            api_url(table),
            headers=supabase_headers("return=representation"),
            data=json.dumps(rows, default=str),
            timeout=60,
        )
    except requests.RequestException as exc:
        st.error("No se pudo conectar con Supabase.")
        st.code(str(exc))
        st.stop()
    if not response.ok:
        stop_supabase_error(response, f"guardar en {table}")
    return response.json()


def sb_patch(table: str, params: dict[str, str], values: dict[str, Any]) -> None:
    try:
        response = requests.patch(
            api_url(table),
            headers=supabase_headers(),
            params=params,
            data=json.dumps(values, default=str),
            timeout=30,
        )
    except requests.RequestException as exc:
        st.error("No se pudo conectar con Supabase.")
        st.code(str(exc))
        st.stop()
    if not response.ok:
        stop_supabase_error(response, f"actualizar {table}")


def sb_delete(table: str, params: dict[str, str]) -> None:
    try:
        response = requests.delete(api_url(table), headers=supabase_headers(), params=params, timeout=30)
    except requests.RequestException as exc:
        st.error("No se pudo conectar con Supabase.")
        st.code(str(exc))
        st.stop()
    if not response.ok:
        stop_supabase_error(response, f"eliminar de {table}")


def delete_advances(advance_ids: list[str]) -> None:
    clean_ids = [item for item in advance_ids if item]
    for start in range(0, len(clean_ids), 100):
        chunk = ",".join(clean_ids[start : start + 100])
        sb_delete("advances", {"id": f"in.({chunk})"})


def is_status_advance(advance: dict[str, Any]) -> bool:
    action = str(advance.get("action") or "").strip().upper()
    return bool(action) and action != "COMENTARIO"


def advance_has_comment(advance: dict[str, Any]) -> bool:
    reason = str(advance.get("reason") or "").strip()
    observation = str(advance.get("observation") or "").strip()
    return bool(reason or observation)


def advance_record_type(advance: dict[str, Any]) -> str:
    if not is_status_advance(advance):
        return "Comentario"
    if advance_has_comment(advance):
        return "Cambio de estado + comentario"
    return "Cambio de estado"


def delete_candidates_preserving_latest_status(
    advances: list[dict[str, Any]],
    candidate_ids: list[str],
) -> tuple[list[str], set[str]]:
    candidate_set = {str(item) for item in candidate_ids if item}
    keep_ids: set[str] = set()
    seen_tasks: set[str] = set()
    for advance in sorted(advances, key=lambda item: item.get("created_at", ""), reverse=True):
        task_id = str(advance.get("task_id") or "")
        advance_id = str(advance.get("id") or "")
        if not task_id or task_id in seen_tasks or not is_status_advance(advance):
            continue
        seen_tasks.add(task_id)
        if advance_id in candidate_set:
            keep_ids.add(advance_id)
    delete_ids = [str(item) for item in candidate_ids if item and str(item) not in keep_ids]
    return delete_ids, keep_ids


def delete_candidates_preserving_latest_records(
    advances: list[dict[str, Any]],
    candidate_ids: list[str],
    delete_status: bool,
    delete_comments: bool,
) -> tuple[list[str], set[str]]:
    candidate_set = {str(item) for item in candidate_ids if item}
    keep_ids: set[str] = set()
    seen_status_tasks: set[str] = set()
    seen_comment_tasks: set[str] = set()
    for advance in sorted(advances, key=lambda item: item.get("created_at", ""), reverse=True):
        task_id = str(advance.get("task_id") or "")
        advance_id = str(advance.get("id") or "")
        if not task_id or not advance_id:
            continue
        if is_status_advance(advance):
            if task_id in seen_status_tasks:
                continue
            seen_status_tasks.add(task_id)
            if advance_id in candidate_set and delete_status:
                keep_ids.add(advance_id)
        else:
            if task_id in seen_comment_tasks:
                continue
            seen_comment_tasks.add(task_id)
            if advance_id in candidate_set and delete_comments:
                keep_ids.add(advance_id)

    delete_ids: list[str] = []
    for advance in advances:
        advance_id = str(advance.get("id") or "")
        if advance_id not in candidate_set or advance_id in keep_ids:
            continue
        if is_status_advance(advance) and delete_status:
            delete_ids.append(advance_id)
        elif not is_status_advance(advance) and delete_comments:
            delete_ids.append(advance_id)
    return delete_ids, keep_ids


def login_screen() -> None:
    app_header("Trabajos programados")
    with st.form("login_form"):
        mode = st.radio("Modo de ingreso", ["Cuadrilla / contratista", "Administrador"], horizontal=True)
        password = st.text_input("Contraseña", type="password")
        submitted = st.form_submit_button("Ingresar", type="primary")
    if submitted:
        if mode.startswith("Administrador"):
            if password == ADMIN_PASSWORD:
                st.session_state.role = "admin"
                st.rerun()
            st.error("Clave administradora incorrecta.")
        else:
            if password == ACCESS_PASSWORD:
                st.session_state.role = "user"
                st.rerun()
            st.error("Contraseña incorrecta.")
    st.stop()


def profile_screen() -> None:
    app_header("Ingreso al programa")
    area = st.selectbox("Area", AREAS, index=0, key="profile_area_select")
    company_options = COMPANIES_BY_AREA[area]
    sector_options = SECTORS_BY_AREA[area]
    is_admin = st.session_state.get("role") == "admin"
    with st.form("profile_form"):
        admin_view = "Programa de tareas"
        if is_admin:
            admin_view = st.radio(
                "Ingresar a",
                ["Programa de tareas", "Avances / registros"],
                horizontal=True,
            )
        company = st.selectbox(
            "Empresa",
            company_options,
            index=0,
            format_func=lambda item: option_label(item, "Todas"),
            key=f"profile_company_{area}",
        )
        sector = st.selectbox(
            "Sector",
            sector_options,
            index=0,
            format_func=lambda item: option_label(item, "Todos"),
            key=f"profile_sector_{area}",
        )
        name_label = "Nombre" if not is_admin else "Nombre (opcional)"
        name = st.text_input(name_label, placeholder="Nombre y apellido / rol")
        if st.form_submit_button("Ingresar", type="primary"):
            if not is_admin and not name.strip():
                st.warning("Ingresar nombre de usuario.")
                st.stop()
            display_name = name.strip() or "Administrador"
            st.session_state.profile = {
                "area": area,
                "company": company,
                "sector": sector,
                "name": display_name,
                "admin_view": admin_view,
            }
            st.rerun()
    st.stop()


def require_session() -> None:
    if "role" not in st.session_state:
        login_screen()
    if "profile" not in st.session_state:
        profile_screen()


FILTER_KEYS = [
    "program_filter",
    "sector_filter",
    "crew_filter",
    "start_filter",
    "end_filter",
    "search_terms_filter",
]


def editor_has_status_changes(editor_state: Any) -> bool:
    """True solo si el editor tiene cambios de Estado. Marcar/desmarcar la
    columna Seleccionar NO se considera un cambio pendiente."""
    if not isinstance(editor_state, dict):
        return False
    for changes in editor_state.get("edited_rows", {}).values():
        if isinstance(changes, dict) and any(key != "Seleccionar" for key in changes):
            return True
    return False


def mark_comment_dirty() -> None:
    """Marca el comentario como pendiente solo cuando el usuario modifica su texto."""
    st.session_state["comment_dirty"] = bool(
        str(st.session_state.get("common_comment_text") or "").strip()
    )


def has_pending_work() -> bool:
    pending = st.session_state.get("pending_state_changes", {})
    comment = str(st.session_state.get("common_comment_text") or "").strip()
    comment_pending = bool(st.session_state.get("comment_dirty", False)) and bool(comment)
    editor_state = st.session_state.get("task_editor", {})
    # La seleccion de tareas (selected_task_ids / check maestro) no cuenta como
    # trabajo pendiente. Un comentario solo cuenta si fue modificado y aun no se guardo.
    return bool(pending) or comment_pending or editor_has_status_changes(editor_state)


def clear_task_selection() -> None:
    """Limpia toda la seleccion de tareas y destilda el check maestro.

    El check maestro se fuerza a False por asignacion explicita (no pop):
    en Streamlit, hacer pop de la key de un widget NO destilda el checkbox
    porque conserva su estado interno; asignarle False si lo destilda."""
    st.session_state.pop("selected_task_ids", None)
    st.session_state["select_all_visible_tasks_filter"] = False
    st.session_state["select_all_visible_tasks_prev"] = False
    st.session_state.pop("task_editor", None)


def clear_pending_work() -> None:
    st.session_state.pop("pending_state_changes", None)
    st.session_state.pop("pending_program_id", None)
    st.session_state.pop("selected_task_ids", None)
    st.session_state["select_all_visible_tasks_filter"] = False
    st.session_state["select_all_visible_tasks_prev"] = False
    st.session_state.pop("task_editor", None)
    st.session_state["comment_dirty"] = False
    st.session_state.clear_comment_text_next = True
    st.session_state.pop("confirm_refresh", None)
    st.session_state.pop("pending_navigation", None)
    st.session_state.pop("filter_change_guard", None)


def current_filter_state(program_id: str) -> dict[str, Any]:
    return {
        "program_id": program_id,
        "program_filter": st.session_state.get("program_filter"),
        "sector_filter": st.session_state.get("sector_filter", ""),
        "crew_filter": st.session_state.get("crew_filter", ""),
        "start_filter": st.session_state.get("start_filter"),
        "end_filter": st.session_state.get("end_filter"),
        "search_terms_filter": tuple(st.session_state.get("search_terms_filter", [])),
    }


def restore_filter_state(state: dict[str, Any] | None) -> None:
    if not state:
        return
    for key in FILTER_KEYS:
        if key in state:
            value = state[key]
            if key == "search_terms_filter" and isinstance(value, tuple):
                value = list(value)
            st.session_state[key] = value


def add_search_filter_from_state() -> None:
    field = str(st.session_state.get("search_field_input") or "Titulo tarea")
    value = str(st.session_state.get("search_value_input") or "").strip()
    if not value:
        return
    label = make_search_filter_label(field, value)
    active = list(st.session_state.get("search_terms_filter", []))
    if normalize(label) not in {normalize(item) for item in active}:
        active.append(label)
    st.session_state.search_terms_filter = active
    st.session_state.search_value_input = ""


def remove_search_filter(label: str) -> None:
    st.session_state.search_terms_filter = [
        item for item in st.session_state.get("search_terms_filter", []) if item != label
    ]


def request_navigation(action: str) -> None:
    if has_pending_work():
        st.session_state.pending_navigation = action
    elif action == "profile":
        clear_task_selection()
        st.session_state.pop("profile", None)
        st.rerun()
    elif action == "logout":
        st.session_state.clear()
        st.rerun()


def apply_navigation(action: str) -> None:
    if action == "profile":
        clear_pending_work()
        st.session_state.pop("profile", None)
        st.rerun()
    if action == "logout":
        st.session_state.clear()
        st.rerun()


def undo_last_change() -> None:
    if st.session_state.get("confirm_refresh") or st.session_state.get("pending_navigation") or st.session_state.get("filter_change_guard"):
        st.session_state.pop("confirm_refresh", None)
        st.session_state.pop("pending_navigation", None)
        st.session_state.pop("filter_change_guard", None)
        clear_task_selection()
        st.rerun()
    if has_pending_work():
        clear_pending_work()
        st.rerun()
    if st.session_state.get("selected_task_ids") or st.session_state.get("select_all_visible_tasks_filter"):
        clear_task_selection()
        st.rerun()
    previous = st.session_state.get("previous_filter_state")
    if previous:
        restore_filter_state(previous)
        st.session_state.last_filter_state = previous
        st.session_state.pop("previous_filter_state", None)
        clear_task_selection()
        st.rerun()


def logout_controls() -> None:
    col1, col2, col3, col4, col5 = st.columns([3.0, 0.9, 0.9, 1, 1])
    profile = st.session_state.profile
    role = "Administrador" if st.session_state.role == "admin" else "Usuario"
    area = canonical_area(profile.get("area") or "GENERACION")
    profile_name = profile.get("name") or role
    col1.caption(
        f"{role}: {profile_name} - {area} - {profile.get('company') or 'Todas'} / {profile.get('sector') or 'Todos'}"
    )
    if col2.button("Deshacer"):
        undo_last_change()
    if col3.button("Actualizar datos"):
        st.session_state.request_refresh = True
    if col4.button("Volver al inicio"):
        request_navigation("profile")
    if col5.button("Cerrar sesion"):
        request_navigation("logout")


def find_column(columns: list[str], candidates: list[str]) -> str | None:
    normalized = {normalize(column): column for column in columns}
    compact = {compact_key(column): column for column in columns}
    for candidate in candidates:
        wanted = normalize(candidate)
        if wanted in normalized:
            return normalized[wanted]
        compact_wanted = compact_key(candidate)
        if compact_wanted in compact:
            return compact[compact_wanted]
    for norm, original in normalized.items():
        if any(normalize(candidate) in norm for candidate in candidates):
            return original
    for norm, original in compact.items():
        if any(compact_key(candidate) in norm for candidate in candidates):
            return original
    return None


def parse_date(value: Any) -> str | None:
    if pd.isna(value) or value == "":
        return None
    try:
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
        if pd.isna(parsed):
            return None
        return parsed.date().isoformat()
    except Exception:
        return None


def duration_days(start: str | None, end: str | None) -> int:
    try:
        start_date = date.fromisoformat(start or end or "")
        end_date = date.fromisoformat(end or start or "")
        return max((end_date - start_date).days + 1, 1)
    except Exception:
        return 1


def row_text(row: pd.Series, column: str | None) -> str:
    if not column:
        return ""
    value = row.get(column, "")
    if pd.isna(value):
        return ""
    return str(value).strip()


def ot_text(value: Any) -> str:
    text = str(value or "").strip()
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
    except Exception:
        pass
    return text


def raw_value(task: dict[str, Any], candidates: list[str]) -> str:
    raw = task.get("raw") or {}
    if not isinstance(raw, dict):
        return ""
    normalized = {normalize(key): value for key, value in raw.items()}
    compact = {compact_key(key): value for key, value in raw.items()}
    for candidate in candidates:
        value = normalized.get(normalize(candidate))
        if value not in (None, ""):
            return str(value).strip()
        value = compact.get(compact_key(candidate))
        if value not in (None, ""):
            return str(value).strip()
    for key, value in normalized.items():
        if value in (None, ""):
            continue
        if any(normalize(candidate) in key for candidate in candidates):
            return str(value).strip()
    for key, value in compact.items():
        if value in (None, ""):
            continue
        if any(compact_key(candidate) in key for candidate in candidates):
            return str(value).strip()
    return ""


def raw_column_name(raw: dict[str, Any], candidates: list[str]) -> str | None:
    normalized = {normalize(column): column for column in raw.keys()}
    compact = {compact_key(column): column for column in raw.keys()}
    for candidate in candidates:
        wanted = normalize(candidate)
        if wanted in normalized:
            return normalized[wanted]
        compact_wanted = compact_key(candidate)
        if compact_wanted in compact:
            return compact[compact_wanted]
    for norm, original in normalized.items():
        if any(normalize(candidate) in norm for candidate in candidates):
            return original
    for norm, original in compact.items():
        if any(compact_key(candidate) in norm for candidate in candidates):
            return original
    return None


def task_title(task: dict[str, Any]) -> str:
    title = raw_value(task, TITLE_COLUMNS)
    if title:
        return title
    return str(task.get("tarea") or "").strip()


def is_distribution_skip_section(value: Any) -> bool:
    return "tareasrelevantes" in compact_key(value)


def section_text(value: Any) -> str:
    text = str(value or "").strip()
    if text.startswith("/") and text.endswith("/"):
        text = text.strip("/").strip()
    return text


def map_excel(df: pd.DataFrame, area: str = "GENERACION") -> list[dict[str, Any]]:
    area = canonical_area(area)
    columns = list(df.columns)
    mapping = {
        "nro_ot": find_column(columns, OT_COLUMNS),
        "tarea": find_column(columns, TITLE_COLUMNS),
        "empresa": find_column(columns, COMPANY_COLUMNS),
        "sector": find_column(columns, SECTOR_COLUMNS),
        "cuadrilla": find_column(columns, CREW_COLUMNS),
        "fecha_inicio": find_column(columns, START_DATE_COLUMNS),
        "fecha_fin": find_column(columns, END_DATE_COLUMNS),
        "estado_programa": find_column(columns, STATUS_COLUMNS),
        "ubicacion_tecnica": find_column(columns, LOCATION_COLUMNS),
        "kks_tag": find_column(columns, KKS_COLUMNS),
    }
    tasks: list[dict[str, Any]] = []
    current_section = ""
    skip_section = False
    for _, row in df.iterrows():
        tarea = row_text(row, mapping["tarea"])
        nro_ot = ot_text(row_text(row, mapping["nro_ot"]))
        cuadrilla = row_text(row, mapping["cuadrilla"])
        estado_programa = row_text(row, mapping["estado_programa"])
        section_candidate = section_text(tarea)
        if area == "DISTRIBUCION":
            section_company, section_sector = distribution_company_sector(section_candidate)
            if is_distribution_skip_section(section_candidate):
                current_section = section_candidate
                skip_section = True
                continue
            is_section_row = bool(section_sector) and not nro_ot and not cuadrilla
            if is_section_row:
                current_section = section_candidate
                skip_section = False
                continue
            if skip_section or not current_section:
                continue
            if not cuadrilla:
                continue
        elif tarea.startswith("/") and tarea.endswith("/") and not nro_ot:
            continue
        if not tarea and not nro_ot:
            continue
        start = parse_date(row.get(mapping["fecha_inicio"])) if mapping["fecha_inicio"] else None
        end = parse_date(row.get(mapping["fecha_fin"])) if mapping["fecha_fin"] else None
        if area == "DISTRIBUCION":
            inferred_company, inferred_sector = distribution_company_sector(current_section)
        else:
            inferred_company, inferred_sector = infer_company_sector(cuadrilla, area)
        empresa = canonical_company(row_text(row, mapping["empresa"])) or inferred_company or canonical_company(cuadrilla)
        sector = canonical_sector(row_text(row, mapping["sector"])) or inferred_sector
        raw = {str(col): (None if pd.isna(row.get(col)) else str(row.get(col))) for col in columns}
        # Guardar la posicion original de la fila dentro del JSON raw para no
        # necesitar una columna adicional en Supabase. "Orden" sigue quedando
        # disponible como alias del numero de OT del Excel.
        raw["_posicion_excel"] = len(tasks)
        digest = hashlib.sha1(json.dumps(raw, sort_keys=True, ensure_ascii=False).encode("utf-8")).hexdigest()
        tasks.append(
            {
                "area": area,
                "row_hash": digest,
                "nro_ot": nro_ot,
                "tarea": tarea,
                "empresa": empresa,
                "sector": sector,
                "cuadrilla": cuadrilla,
                "fecha_inicio": start,
                "fecha_fin": end,
                "duracion": duration_days(start, end),
                "estado_programa": estado_programa,
                "ubicacion_tecnica": row_text(row, mapping["ubicacion_tecnica"]),
                "kks_tag": row_text(row, mapping["kks_tag"]),
                "raw": raw,
            }
        )
    return tasks


def list_programs(active_only: bool = True) -> list[dict[str, Any]]:
    params = {"select": "*", "order": "uploaded_at.desc"}
    if active_only:
        params["active"] = "eq.true"
    return sb_get("programs", params)


def load_tasks(program_id: str) -> list[dict[str, Any]]:
    tasks = sb_get("tasks", {"select": "*", "program_id": f"eq.{program_id}"})

    def excel_position(task: dict[str, Any]) -> int:
        raw = task.get("raw") or {}
        if not isinstance(raw, dict):
            return 10**9
        try:
            return int(raw.get("_posicion_excel", 10**9))
        except (TypeError, ValueError):
            return 10**9

    # Los programas nuevos respetan la posicion original del Excel guardada
    # dentro de raw. Los programas anteriores, que no tienen _posicion_excel,
    # siguen funcionando y se ordenan como respaldo por fecha de inicio y OT.
    tasks.sort(
        key=lambda t: (
            excel_position(t),
            str(t.get("fecha_inicio") or ""),
            str(t.get("nro_ot") or ""),
        )
    )
    return tasks


def load_advances(program_id: str) -> list[dict[str, Any]]:
    return sb_get("advances", {"select": "*", "program_id": f"eq.{program_id}", "order": "created_at.desc"})


def load_crew_status(area: str) -> list[dict[str, Any]]:
    return sb_get("crew_status", {"select": "*", "area": f"eq.{area}", "order": "created_at.desc"})


def latest_crew_status_by_crew(records: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for record in sorted(records, key=lambda item: item.get("created_at", ""), reverse=True):
        key = normalize_crew(record.get("cuadrilla"))
        if key and key not in latest:
            latest[key] = record
    return latest


def filter_crew_status(records: list[dict[str, Any]], company: str, sector: str) -> list[dict[str, Any]]:
    company = scope_value(company)
    sector = scope_value(sector)
    filtered = []
    for record in records:
        if company and normalize(record.get("empresa")) != normalize(company):
            continue
        if sector and normalize(record.get("sector")) != normalize(sector):
            continue
        filtered.append(record)
    return filtered


def save_crew_status(
    area: str,
    cuadrilla: str,
    estado_operativo: str,
    integrantes: str,
    movil: str,
    tetra: str,
    empresa: str,
    sector: str,
) -> None:
    profile = st.session_state.profile
    row = {
        "area": area,
        "cuadrilla": cuadrilla,
        "estado_operativo": estado_operativo,
        "integrantes": integrantes,
        "movil": movil,
        "tetra": tetra,
        "empresa": empresa,
        "sector": sector,
        "reporter_name": profile["name"],
        "reporter_company": profile.get("company") or "",
        "reporter_sector": profile.get("sector") or "",
    }
    sb_insert("crew_status", row)


def save_crew_status_entries(area: str, entries: list[dict[str, str]], empresa: str, sector: str) -> None:
    profile = st.session_state.profile
    rows = [
        {
            "area": area,
            "cuadrilla": entry["cuadrilla"],
            "estado_operativo": entry["estado_operativo"],
            "integrantes": entry["integrantes"],
            "movil": entry["movil"],
            "tetra": entry["tetra"],
            "empresa": entry.get("empresa") or empresa,
            "sector": entry.get("sector") or sector,
            "reporter_name": profile["name"],
            "reporter_company": profile.get("company") or "",
            "reporter_sector": profile.get("sector") or "",
        }
        for entry in entries
    ]
    sb_insert("crew_status", rows)


def read_excel_with_split_header(uploaded_file: Any) -> pd.DataFrame:
    """Lee un Excel tolerando encabezados partidos en 2 filas por celdas
    combinadas (ej.: una columna ancha fusionada empuja las siguientes
    columnas a una segunda fila de titulos)."""
    raw = pd.read_excel(uploaded_file, header=None)
    if len(raw) < 2:
        return pd.read_excel(uploaded_file)
    row0 = raw.iloc[0]
    row1 = raw.iloc[1]
    row0_blank = row0.isna()
    row1_blank = row1.isna()
    is_split_header = bool((row0_blank != row1_blank).all()) and not (
        bool(row0_blank.all()) or bool(row1_blank.all())
    )
    if is_split_header:
        header = [row0.iloc[i] if pd.notna(row0.iloc[i]) else row1.iloc[i] for i in range(len(row0))]
        data = raw.iloc[2:].reset_index(drop=True)
    else:
        header = row0.tolist()
        data = raw.iloc[1:].reset_index(drop=True)
    data.columns = header
    return data


def read_crew_status_table(uploaded_file: Any) -> pd.DataFrame | None:
    name = str(getattr(uploaded_file, "name", "") or "").lower()
    if not name.endswith(".pdf"):
        return read_excel_with_split_header(uploaded_file)
    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                if not table or len(table) < 2:
                    continue
                header = [str(cell or "").replace("\n", " ").strip() for cell in table[0]]
                body = table[1:]
                return pd.DataFrame(body, columns=header)
    return None


def blank_if_na(value: str) -> str:
    return "" if normalize(value) in CREW_STATUS_NA_VALUES else value


def map_crew_status_dataframe(df: pd.DataFrame) -> list[dict[str, str]]:
    columns = list(df.columns)
    mapping = {
        "cuadrilla": find_column(columns, CREW_STATUS_CUADRILLA_COLUMNS),
        "movil": find_column(columns, CREW_STATUS_MOVIL_COLUMNS),
        "tetra": find_column(columns, CREW_STATUS_TETRA_COLUMNS),
        "encargado": find_column(columns, CREW_STATUS_ENCARGADO_COLUMNS),
        "operario": find_column(columns, CREW_STATUS_OPERARIO_COLUMNS),
        "integrantes": find_column(columns, CREW_STATUS_INTEGRANTES_COLUMNS),
        "disponibilidad": find_column(columns, CREW_STATUS_DISPONIBILIDAD_COLUMNS),
    }
    has_disponibilidad = mapping["disponibilidad"] is not None
    seen_crews: set[str] = set()
    entries: list[dict[str, str]] = []
    for _, row in df.iterrows():
        cuadrilla = ot_text(row_text(row, mapping["cuadrilla"]))
        if not cuadrilla:
            continue
        crew_key = normalize_crew(cuadrilla)
        if crew_key in seen_crews:
            continue
        seen_crews.add(crew_key)
        encargado = row_text(row, mapping["encargado"])
        operario = row_text(row, mapping["operario"])
        integrantes = " / ".join(part for part in [encargado, operario] if part)
        if not integrantes:
            integrantes = blank_if_na(row_text(row, mapping["integrantes"]))
        if has_disponibilidad:
            disponibilidad = row_text(row, mapping["disponibilidad"])
            estado_operativo = "Operativa" if normalize(disponibilidad) == "disponible" else "No operativa"
        else:
            estado_operativo = "Operativa"
        entries.append(
            {
                "cuadrilla": cuadrilla,
                "estado_operativo": estado_operativo,
                "integrantes": integrantes,
                "movil": ot_text(blank_if_na(row_text(row, mapping["movil"]))),
                "tetra": ot_text(blank_if_na(row_text(row, mapping["tetra"]))),
            }
        )
    return entries


def crew_status_row_style(row: pd.Series) -> list[str]:
    status = str(row.get("Estado operativo") or "").strip().lower()
    style = (
        "background-color: rgba(36, 161, 72, 0.30); color: #f2fff5; font-weight: 700;"
        if status == "operativa"
        else "background-color: rgba(248, 81, 73, 0.30); color: #fff0ef; font-weight: 700;"
    )
    return [style for _ in row]


def delete_crew_status_records(record_ids: list[str]) -> None:
    clean_ids = [item for item in record_ids if item]
    for start in range(0, len(clean_ids), 100):
        chunk = ",".join(clean_ids[start : start + 100])
        sb_delete("crew_status", {"id": f"in.({chunk})"})


def crew_status_rows_for_display(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "_record_id": record.get("id", ""),
            "Cuadrilla": record.get("cuadrilla", ""),
            "Empresa": record.get("empresa") or "",
            "Sector": record.get("sector") or "",
            "Estado operativo": record.get("estado_operativo", ""),
            "Integrantes": record.get("integrantes") or "",
            "Movil": record.get("movil") or "",
            "Tetra": record.get("tetra") or "",
            "Informado por": record.get("reporter_name", ""),
            "Fecha": advance_date(record.get("created_at")),
            "Hora": advance_time(record.get("created_at")),
        }
        for record in records
    ]


def crew_status_export(records: list[dict[str, Any]]) -> bytes:
    rows = [
        {key: value for key, value in row.items() if key != "_record_id"}
        for row in crew_status_rows_for_display(records)
    ]
    return write_excel(rows, sheet_name="Info cuadrillas")


def render_crew_status_section(
    area: str,
    crew_options: list[str],
    filter_company: str,
    filter_sector: str,
    crew_lookup: dict[str, tuple[str, str]] | None = None,
) -> None:
    crew_lookup = crew_lookup or {}
    with st.expander("Info de cuadrillas", expanded=False):
        st.caption(
            "Estado operativo por cuadrilla, independiente de las tareas del programa. "
            "Se filtra por la misma Empresa/Sector que el tablero de tareas. Si la "
            "cuadrilla ya aparece en el programa activo, la Empresa/Sector se detecta "
            "sola a partir de esos datos."
        )
        profile = st.session_state.profile
        company_options = COMPANIES_BY_AREA[area]
        sector_options = SECTORS_BY_AREA[area]
        profile_company = profile.get("company") or ""
        profile_sector = profile.get("sector") or ""
        company_index = company_options.index(profile_company) if profile_company in company_options else 0
        sector_index = sector_options.index(profile_sector) if profile_sector in sector_options else 0
        is_distribution = area == "DISTRIBUCION"

        records = filter_crew_status(load_crew_status(area), filter_company, filter_sector)
        latest = latest_crew_status_by_crew(records)

        manual_entry_enabled = st.checkbox(
            "Agregar/editar una cuadrilla manualmente",
            key="crew_status_manual_toggle",
        )
        if manual_entry_enabled:
            with st.form("crew_status_form", clear_on_submit=True):
                crew_col1, crew_col2 = st.columns([1.0, 1.3])
                suggested_crew = crew_col1.selectbox(
                    "Cuadrilla existente",
                    [""] + crew_options,
                    format_func=lambda item: option_label(item, "Elegir / escribir nueva"),
                    key="crew_status_suggested",
                )
                manual_crew = crew_col2.text_input(
                    "Cuadrilla",
                    placeholder="Ej.: 555, 556A, 720",
                    key="crew_status_manual",
                )
                if is_distribution:
                    manual_sector = st.selectbox(
                        "Sector",
                        sector_options,
                        index=sector_index,
                        format_func=lambda item: option_label(item, "Todos"),
                        key="crew_status_sector",
                    )
                    manual_company_auto = distribution_sector_company(manual_sector)
                    if manual_company_auto:
                        st.caption(f"Empresa segun sector: **{manual_company_auto}**")
                        manual_company = manual_company_auto
                    else:
                        manual_company = st.selectbox(
                            "Empresa",
                            company_options,
                            index=company_index,
                            format_func=lambda item: option_label(item, "Todas"),
                            key="crew_status_company",
                        )
                else:
                    manual_company_col, manual_sector_col = st.columns(2)
                    manual_company = manual_company_col.selectbox(
                        "Empresa",
                        company_options,
                        index=company_index,
                        format_func=lambda item: option_label(item, "Todas"),
                        key="crew_status_company",
                    )
                    manual_sector = manual_sector_col.selectbox(
                        "Sector",
                        sector_options,
                        index=sector_index,
                        format_func=lambda item: option_label(item, "Todos"),
                        key="crew_status_sector",
                    )
                estado_operativo = st.radio(
                    "Estado operativo",
                    CREW_OPERATIONAL_STATES,
                    horizontal=True,
                    key="crew_status_estado",
                )
                info_col1, info_col2, info_col3 = st.columns(3)
                integrantes = info_col1.text_input("Integrantes", key="crew_status_integrantes")
                movil = info_col2.text_input("Movil", key="crew_status_movil")
                tetra = info_col3.text_input("Tetra", key="crew_status_tetra")
                submitted = st.form_submit_button("Guardar info cuadrilla", type="primary")

            if submitted:
                crew_value = (manual_crew or suggested_crew).strip()
                if not crew_value:
                    st.warning("Elegi o escribi una cuadrilla.")
                else:
                    detected = crew_lookup.get(normalize_crew(crew_value))
                    final_company = (detected[0] if detected and detected[0] else "") or manual_company
                    final_sector = (detected[1] if detected and detected[1] else "") or manual_sector
                    if not scope_value(final_company):
                        st.warning(
                            "Elegi una Empresa especifica (no 'Todas') para que la cuadrilla "
                            "sea visible para los usuarios de esa empresa."
                        )
                    else:
                        if detected:
                            st.caption(
                                f"Empresa/Sector detectados del programa activo: "
                                f"{final_company} / {final_sector or 'sin sector'}."
                            )
                        save_crew_status(
                            area,
                            crew_value,
                            estado_operativo,
                            integrantes.strip(),
                            movil.strip(),
                            tetra.strip(),
                            final_company,
                            final_sector,
                        )
                        st.success(f"Info de cuadrilla guardada para {crew_value}.")
                        st.rerun()

        st.markdown("**Importar cuadrillas desde Excel o PDF**")
        st.caption(
            "Reconoce columnas Cuadrilla, Encargado, Operario, Interno/Movil, Tetra y "
            "Disponibilidad. Si hay varias filas de la misma cuadrilla, se toma la primera. "
            "'Disponible' se carga como Operativa, cualquier otro valor como No operativa. "
            "El PDF depende de como este armado el reporte; si no lo puede leer, probar con el Excel. "
            "Empresa es el dato obligatorio del lote (quien informa el archivo); el Sector de "
            "cada cuadrilla se detecta solo si ya aparece en el programa activo. Para las que "
            "no se detecten, se usa el Sector por defecto que elijas abajo (opcional)."
        )
        import_company = st.selectbox(
            "Empresa de estas cuadrillas",
            company_options,
            index=company_index,
            format_func=lambda item: option_label(item, "Todas"),
            key="crew_status_import_company",
        )
        import_sector = st.selectbox(
            "Sector por defecto (para cuadrillas no detectadas)",
            sector_options,
            index=sector_index,
            format_func=lambda item: option_label(item, "Todos"),
            key="crew_status_import_sector",
        )
        uploaded_crew_file = st.file_uploader(
            "Archivo de ubicacion de cuadrillas",
            type=["xlsx", "xls", "pdf"],
            key="crew_status_import_file",
        )
        import_company_missing = not scope_value(import_company)
        if import_company_missing:
            st.warning(
                "Elegi arriba la Empresa que informa estas cuadrillas (no 'Todas') antes "
                "de importar: si queda en 'Todas' las cuadrillas se guardan sin empresa y "
                "no aparecen cuando un usuario filtra por su empresa o sector."
            )
        if st.button(
            "Importar cuadrillas",
            disabled=uploaded_crew_file is None or import_company_missing,
            key="crew_status_import_btn",
        ):
            with st.spinner("Importando archivo..."):
                import_df = read_crew_status_table(uploaded_crew_file)
            if import_df is None or import_df.empty:
                st.error("No pude leer una tabla de cuadrillas en ese archivo. Probar con el Excel.")
            else:
                entries = map_crew_status_dataframe(import_df)
                if not entries:
                    st.error("No encontre cuadrillas validas en el archivo.")
                else:
                    undetected: list[str] = []
                    for entry in entries:
                        detected = crew_lookup.get(normalize_crew(entry["cuadrilla"]))
                        detected_sector = detected[1] if detected else ""
                        if detected_sector:
                            entry["sector"] = detected_sector
                            entry["empresa"] = (
                                distribution_sector_company(detected_sector)
                                if is_distribution
                                else (detected[0] if detected and detected[0] else "")
                            ) or import_company
                        else:
                            undetected.append(entry["cuadrilla"])
                            entry["sector"] = import_sector
                            entry["empresa"] = import_company
                    save_crew_status_entries(area, entries, import_company, import_sector)
                    st.success(f"Info cargada para {len(entries)} cuadrilla(s).")
                    if undetected:
                        st.info(
                            "No se detecto el Sector en el programa activo para: "
                            f"{', '.join(undetected)}. Se cargaron con Empresa {import_company} / "
                            f"Sector {import_sector or 'sin sector'}."
                        )
                    st.rerun()

        if not records:
            st.info("Todavia no hay info de cuadrillas cargada para este filtro de area/empresa/sector.")
            return

        view_choice = st.selectbox(
            "Vista",
            ["Estado actual por cuadrilla", "Log completo"],
            key="crew_status_view",
        )
        source_records = (
            sorted(latest.values(), key=lambda item: normalize(item.get("cuadrilla")))
            if view_choice == "Estado actual por cuadrilla"
            else records
        )
        display_rows = crew_status_rows_for_display(source_records)
        display_df = pd.DataFrame(
            [{key: value for key, value in row.items() if key != "_record_id"} for row in display_rows]
        )
        st.dataframe(
            display_df.style.apply(crew_status_row_style, axis=1),
            hide_index=True,
            use_container_width=False,
        )

        st.download_button(
            "Exportar info cuadrillas Excel",
            data=crew_status_export(source_records),
            file_name=f"info_cuadrillas_{local_today().isoformat()}.xlsx",
            key="export_crew_status",
        )

        if st.session_state.role == "admin":
            with st.expander("Administracion de registros de cuadrillas", expanded=False):
                visible_ids = tuple(str(row.get("_record_id", "")) for row in display_rows)
                snapshot_key = f"crew_status_visible_snapshot_{view_choice}"
                delete_editor_key = f"delete_crew_status_editor_{view_choice}"
                if st.session_state.get(snapshot_key) != visible_ids:
                    st.session_state.pop("select_all_visible_crew_status", None)
                    st.session_state.pop(delete_editor_key, None)
                    st.session_state[snapshot_key] = visible_ids
                select_all_visible = st.checkbox(
                    "Seleccionar todo visible",
                    key="select_all_visible_crew_status",
                )
                delete_df = pd.DataFrame([{**{"Eliminar": False}, **row} for row in display_rows])
                if select_all_visible:
                    delete_df["Eliminar"] = True
                edited_delete = st.data_editor(
                    delete_df,
                    hide_index=True,
                    use_container_width=False,
                    disabled=[column for column in delete_df.columns if column != "Eliminar"],
                    column_config={
                        "_record_id": None,
                        "Eliminar": st.column_config.CheckboxColumn("Eliminar", width=76),
                    },
                    key=delete_editor_key,
                )
                selected_delete_ids = (
                    edited_delete.loc[edited_delete["Eliminar"] == True, "_record_id"].tolist()
                    if not edited_delete.empty
                    else []
                )
                if st.button(
                    "Borrar seleccionados",
                    disabled=not selected_delete_ids,
                    key="delete_crew_status_btn",
                ):
                    delete_crew_status_records(selected_delete_ids)
                    st.session_state.pop(delete_editor_key, None)
                    st.session_state.pop("select_all_visible_crew_status", None)
                    st.success(f"{len(selected_delete_ids)} registro(s) eliminado(s).")
                    st.rerun()


def admin_panel() -> None:
    if st.session_state.role != "admin":
        return
    with st.expander("Administracion - cargar programa Excel", expanded=False):
        default_area = current_area()
        upload_area = st.selectbox("Area del programa", AREAS, index=AREAS.index(default_area))
        uploaded = st.file_uploader("Excel del programa", type=["xlsx", "xls"])
        program_name = st.text_input("Nombre del programa", value=f"Programa {local_today().isoformat()}")
        replace_active = st.checkbox("Dejar este como unico programa activo", value=True)
        if st.button("Publicar programa", type="primary", disabled=uploaded is None):
            with st.spinner("Importando Excel..."):
                df = pd.read_excel(uploaded)
                tasks = map_excel(df, upload_area)
                if not tasks:
                    st.error("No encontre tareas validas en el Excel.")
                    return
                if replace_active:
                    for program in list_programs(active_only=True):
                        if program_area(program) == upload_area:
                            sb_patch("programs", {"id": f"eq.{program['id']}"}, {"active": False})
                program = sb_insert(
                    "programs",
                    {
                        "name": program_name.strip() or uploaded.name,
                        "area": upload_area,
                        "source_filename": uploaded.name,
                        "uploaded_by": st.session_state.profile["name"],
                        "active": True,
                    },
                )[0]
                for task in tasks:
                    task["program_id"] = program["id"]
                for start in range(0, len(tasks), 500):
                    sb_insert("tasks", tasks[start : start + 500])
            st.success(f"Programa publicado con {len(tasks)} tarea(s).")
            st.rerun()

        programs = list_programs(active_only=False)
        if programs:
            st.divider()
            st.subheader("Agregar tarea manual al programa")
            st.caption(
                "Permite incorporar una tarea nueva sin volver a importar el Excel. "
                "La tarea queda vinculada al programa activo que elijas."
            )

            active_manual_programs = [
                program
                for program in list_programs(active_only=True)
                if program_area(program) == current_area()
            ]

            if not active_manual_programs:
                st.info("No hay programas activos disponibles para agregar una tarea manual.")
            else:
                current_program_id = ""
                current_program_value = st.session_state.get("program_filter")
                if isinstance(current_program_value, dict):
                    current_program_id = str(current_program_value.get("id") or "")

                manual_program_index = next(
                    (
                        idx
                        for idx, item in enumerate(active_manual_programs)
                        if str(item.get("id") or "") == current_program_id
                    ),
                    0,
                )

                manual_program = st.selectbox(
                    "Programa activo donde se agregara la tarea",
                    active_manual_programs,
                    index=manual_program_index,
                    format_func=lambda item: f"{program_area(item)} - {item['name']}",
                    key="manual_task_program",
                )

                manual_area = program_area(manual_program)
                existing_tasks = load_tasks(str(manual_program["id"]))

                existing_crews = sorted(
                    {
                        str(task.get("cuadrilla") or "").strip()
                        for task in existing_tasks
                        if str(task.get("cuadrilla") or "").strip()
                    }
                )

                st.info(
                    f"Destino seleccionado: {manual_area} / Programa: {manual_program['name']}"
                )

                with st.form("manual_task_form", clear_on_submit=True):
                    manual_description = st.text_input("Descripcion de la tarea *")

                    date_col1, date_col2 = st.columns(2)
                    manual_start = date_col1.date_input(
                        "Fecha inicio *",
                        value=local_today(),
                        format="DD/MM/YYYY",
                    )
                    manual_end = date_col2.date_input(
                        "Fecha fin *",
                        value=local_today(),
                        format="DD/MM/YYYY",
                    )

                    crew_col1, crew_col2 = st.columns([1.0, 1.25])
                    suggested_crew = crew_col1.selectbox(
                        "Cuadrilla existente",
                        [""] + existing_crews,
                        format_func=lambda item: option_label(item, "Elegir / escribir nueva"),
                    )
                    manual_crew = crew_col2.text_input(
                        "Cuadrilla",
                        placeholder="Ej.: 555, 556A, 720",
                    )

                    ot_col, pte_col = st.columns(2)
                    manual_ot = ot_col.text_input("OT *", placeholder="Numero de OT")
                    manual_pte = pte_col.text_input("PTE", placeholder="PTE")

                    location_col, kks_col = st.columns([1.6, 1.0])
                    manual_location = location_col.text_input(
                        "Ubicacion tecnica",
                        placeholder="Ubicacion tecnica",
                    )
                    manual_kks = kks_col.text_input(
                        "KKS-TAG",
                        placeholder="KKS / TAG",
                    )

                    manual_submit = st.form_submit_button(
                        "Agregar tarea al programa",
                        type="primary",
                        use_container_width=True,
                    )

                if manual_submit:
                    description = manual_description.strip()
                    crew_value = manual_crew.strip() or suggested_crew.strip()
                    ot_value = ot_text(manual_ot)

                    if not description:
                        st.error("Completar la descripcion de la tarea.")
                    elif not crew_value:
                        st.error("Completar o elegir la cuadrilla.")
                    elif not ot_value:
                        st.error("Completar la OT.")
                    elif manual_end < manual_start:
                        st.error("La fecha fin no puede ser anterior a la fecha inicio.")
                    else:
                        duplicate = next(
                            (
                                task
                                for task in existing_tasks
                                if normalize(ot_text(task.get("nro_ot"))) == normalize(ot_value)
                            ),
                            None,
                        )

                        if duplicate:
                            st.error(
                                "Esa OT ya existe en el programa seleccionado. "
                                f"Cuadrilla: {duplicate.get('cuadrilla') or '-'} / "
                                f"Tarea: {task_title(duplicate) or '-'}"
                            )
                        else:
                            # Prioridad para empresa/sector: 1) otra tarea del mismo
                            # programa que ya tenga esa misma cuadrilla (dato real
                            # cargado del Excel), 2) tabla de cuadrillas conocidas,
                            # 3) empresa/sector del perfil de quien esta cargando.
                            crew_match = next(
                                (
                                    task
                                    for task in existing_tasks
                                    if normalize_crew(task.get("cuadrilla")) == normalize_crew(crew_value)
                                ),
                                None,
                            )
                            company_from_crew = effective_company(crew_match) if crew_match else ""
                            sector_from_crew = effective_sector(crew_match) if crew_match else ""

                            inferred_company, inferred_sector = infer_company_sector(
                                crew_value,
                                manual_area,
                            )

                            profile = st.session_state.get("profile", {})
                            manual_company = company_from_crew or inferred_company or canonical_company(
                                profile.get("company")
                            )
                            manual_sector = sector_from_crew or inferred_sector or canonical_sector(
                                profile.get("sector")
                            )

                            positions = []
                            for task in existing_tasks:
                                raw = task.get("raw") or {}
                                if not isinstance(raw, dict):
                                    continue
                                try:
                                    positions.append(int(raw.get("_posicion_excel")))
                                except (TypeError, ValueError):
                                    pass
                            next_position = (
                                max(positions) + 1 if positions else len(existing_tasks)
                            )

                            manual_raw = {
                                "_origen": "MANUAL",
                                "_posicion_excel": next_position,
                                "PTE": manual_pte.strip(),
                                "Descripcion": description,
                                "OT": ot_value,
                                "Cuadrilla": crew_value,
                                "Fecha inicio": manual_start.isoformat(),
                                "Fecha fin": manual_end.isoformat(),
                                "Ubicacion tecnica": manual_location.strip(),
                                "KKS-TAG": manual_kks.strip(),
                            }

                            digest_source = {
                                "program_id": str(manual_program["id"]),
                                "manual": manual_raw,
                                "created_at": datetime.now(timezone.utc).isoformat(),
                            }
                            digest = hashlib.sha1(
                                json.dumps(
                                    digest_source,
                                    sort_keys=True,
                                    ensure_ascii=False,
                                ).encode("utf-8")
                            ).hexdigest()

                            task_row = {
                                "program_id": str(manual_program["id"]),
                                "area": manual_area,
                                "row_hash": digest,
                                "nro_ot": ot_value,
                                "tarea": description,
                                "empresa": manual_company,
                                "sector": manual_sector,
                                "cuadrilla": crew_value,
                                "fecha_inicio": manual_start.isoformat(),
                                "fecha_fin": manual_end.isoformat(),
                                "duracion": duration_days(
                                    manual_start.isoformat(),
                                    manual_end.isoformat(),
                                ),
                                "estado_programa": "SIN AVANCE",
                                "ubicacion_tecnica": manual_location.strip(),
                                "kks_tag": manual_kks.strip(),
                                "raw": manual_raw,
                            }

                            sb_insert("tasks", task_row)

                            company_label = manual_company or "Sin empresa detectada"
                            sector_label = manual_sector or "Sin sector detectado"
                            st.success(
                                f"OT {ot_value} agregada correctamente a "
                                f"{manual_program['name']} / {manual_area} / "
                                f"{company_label} / {sector_label} / Cuadrilla {crew_value}."
                            )
                            st.rerun()



def selected_crews(value: Any) -> set[str]:
    if not value:
        return set()
    if isinstance(value, (list, tuple, set)):
        return {normalize_crew(item) for item in value if str(item or "").strip()}
    return {normalize_crew(value)}


def search_terms(value: Any) -> list[str]:
    if not value:
        return []
    if isinstance(value, (list, tuple, set)):
        raw_terms = [str(item or "").strip() for item in value]
    else:
        raw_terms = [str(value or "").strip()]
    terms = []
    seen = set()
    for item in raw_terms:
        normalized = normalize(item)
        if normalized and normalized not in seen:
            terms.append(normalized)
            seen.add(normalized)
    return terms


def make_search_filter_label(field: str, value: str) -> str:
    return f"{field}: {value.strip()}"


def parse_search_filter_label(label: Any) -> tuple[str, str]:
    text = str(label or "").strip()
    if ":" not in text:
        return "Todos los campos", text
    field, value = text.split(":", 1)
    field = field.strip()
    value = value.strip()
    if field not in SEARCH_FIELD_OPTIONS:
        return "Todos los campos", value or text
    return field, value


def task_search_text(task: dict[str, Any], field: str) -> str:
    if field in SEARCH_FIELD_OPTIONS:
        keys = SEARCH_FIELD_OPTIONS[field]
    else:
        keys = ["nro_ot", "tarea", "cuadrilla", "ubicacion_tecnica", "kks_tag"]
    text = " ".join(str(task.get(key) or "") for key in keys)
    if field in {"Todos los campos", "Titulo tarea"}:
        text = f"{text} {task_title(task)}"
    return normalize(text)


def apply_filters(
    tasks: list[dict[str, Any]],
    company: str,
    sector: str,
    crew: Any,
    start: Any,
    end: Any,
    text: Any,
) -> list[dict[str, Any]]:
    output = []
    company = scope_value(company)
    sector = scope_value(sector)
    crew_values = selected_crews(crew)
    start_iso = start.isoformat() if start else ""
    end_iso = end.isoformat() if end else ""
    search_filters = [parse_search_filter_label(item) for item in (text or [])] if isinstance(text, (list, tuple, set)) else []
    if not search_filters and text:
        search_filters = [("Todos los campos", str(text))]
    search_filters = [(field, value) for field, value in search_filters if normalize(value)]
    for task in tasks:
        if company:
            if normalize(company) == "otra":
                if not is_other_company_scope(task):
                    continue
            elif normalize(effective_company(task)) != normalize(company):
                continue
        if sector:
            if normalize(sector) == "otros":
                if not is_other_sector_scope(task):
                    continue
            elif normalize(effective_sector(task)) != normalize(sector):
                continue
        if crew_values and normalize_crew(task.get("cuadrilla")) not in crew_values:
            continue
        if start_iso or end_iso:
            task_start = str(task.get("fecha_inicio") or "").strip()
            task_end = str(task.get("fecha_fin") or "").strip()
            if task_start and not task_end:
                task_end = task_start
            elif task_end and not task_start:
                task_start = task_end
            if not task_start and not task_end:
                continue
            if start_iso and end_iso:
                if task_start < start_iso or task_end > end_iso:
                    continue
            elif start_iso:
                if task_start != start_iso:
                    continue
            elif end_iso:
                if task_start > end_iso or task_end < end_iso:
                    continue
        if search_filters:
            filters_by_field: dict[str, list[str]] = {}
            for field, value in search_filters:
                filters_by_field.setdefault(field, []).append(value)
            if not all(
                any(normalize(value) in task_search_text(task, field) for value in values)
                for field, values in filters_by_field.items()
            ):
                continue
        output.append(task)
    return output


def latest_status_by_task(advances: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    latest: dict[str, dict[str, Any]] = {}
    for advance in sorted(advances, key=lambda item: item.get("created_at", ""), reverse=True):
        if advance["task_id"] not in latest and advance.get("action") != "COMENTARIO":
            latest[advance["task_id"]] = advance
    return latest


def effective_task_status(task: dict[str, Any], latest: dict[str, dict[str, Any]]) -> str:
    """Estado efectivo de la tarea: prioriza el ultimo avance guardado en la app
    y, si no hay, usa el estado que trae el programa (Excel). Normaliza a
    mayusculas para tolerar distinto formato en el Excel."""
    advance = latest.get(task["id"], {})
    status = str(advance.get("action") or task.get("estado_programa") or "SIN AVANCE").strip().upper()
    if status not in STATE_ACTIONS:
        status = "SIN AVANCE"
    return status


def task_date_by_mode(task: dict[str, Any], mode: str, manual_value: date | None, field: str) -> str | None:
    if mode == DATE_MODE_PROGRAM:
        source = task.get(field)
        if not source:
            candidates = START_DATE_COLUMNS if field == "fecha_inicio" else END_DATE_COLUMNS
            source = raw_value(task, candidates)
        return parse_date(source)
    if mode == DATE_MODE_MANUAL:
        return manual_value.isoformat() if manual_value else None
    return local_today().isoformat()


def advance_task_start_date(advance: dict[str, Any]) -> str:
    return format_date(advance.get("fecha_inicio_tarea") or "")


def advance_task_finish_date(advance: dict[str, Any]) -> str:
    return format_date(advance.get("fecha_finalizacion_tarea") or "")


def original_start_date(task: dict[str, Any]) -> str:
    return format_date(task.get("fecha_inicio")) or format_date(raw_value(task, START_DATE_COLUMNS))


def original_end_date(task: dict[str, Any]) -> str:
    return format_date(task.get("fecha_fin")) or format_date(raw_value(task, END_DATE_COLUMNS))


def advance_operational_date(advance: dict[str, Any], fallback_created: bool = True) -> str:
    action = str(advance.get("action") or "").strip().upper()
    if action == "EN CURSO":
        value = advance_task_start_date(advance)
        if value:
            return value
    if action == "COMPLETADO":
        value = advance_task_finish_date(advance)
        if value:
            return value
    return advance_date(advance.get("created_at")) if fallback_created else ""


def task_dataframe(tasks: list[dict[str, Any]], latest: dict[str, dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for task in tasks:
        advance = latest.get(task["id"], {})
        status = effective_task_status(task, latest)
        display_start = format_date(task.get("fecha_inicio"))
        if status == "EN CURSO":
            display_start = advance_task_start_date(advance) or display_start
        rows.append(
            {
                "Seleccionar": False,
                "Estado": status,
                "Titulo tarea": task_title(task),
                "Fecha inicio": display_start,
                "Duracion": f"{task.get('duracion') or 1} dia(s)",
                "Cuadrilla": task.get("cuadrilla") or "",
                "OT": ot_text(task.get("nro_ot")),
                "Ubicacion tecnica": task.get("ubicacion_tecnica") or "",
                "PTE": ot_text(raw_value(task, ["PTE", "pte"])),
                "KKS/TAG": task.get("kks_tag") or "",
                "_task_id": task["id"],
                "_estado_original": status,
            }
        )
    return pd.DataFrame(rows)


def task_status_style(row: pd.Series) -> list[str]:
    status = str(row.get("Estado") or "").strip().upper()
    row_style = STATE_ROW_STYLES.get(status, "")
    return [row_style for _ in row]


def record_status_style(row: pd.Series) -> list[str]:
    state_column = "Estado final" if "Estado final" in row.index else "Estado"
    status = str(row.get(state_column) or "").strip().upper()
    return [STATE_ROW_STYLES.get(status, "") for _ in row]


def task_table_column_config(compact: bool = False) -> dict[str, Any]:
    config = {
        "Seleccionar": st.column_config.CheckboxColumn("Sel.", width=42 if compact else 56),
        "Estado": st.column_config.SelectboxColumn("Estado", options=STATE_ACTIONS, required=True, width=100 if compact else 108),
        "Titulo tarea": st.column_config.TextColumn("Titulo tarea", width=280 if compact else 390),
        "Fecha inicio": st.column_config.TextColumn("Fecha inicio", width=90 if compact else 104),
        "Duracion": st.column_config.TextColumn("Duracion", width=82),
        "Cuadrilla": st.column_config.TextColumn("Cuadrilla", width=76 if compact else 82),
        "OT": st.column_config.TextColumn("OT", width=80 if compact else 92),
        "Ubicacion tecnica": st.column_config.TextColumn("Ubicacion tecnica", width=230 if compact else 290),
        "PTE": st.column_config.TextColumn("PTE", width=105),
        "KKS/TAG": st.column_config.TextColumn("KKS/TAG", width=110),
    }
    if compact:
        # En vista compacta (pensada para celular) se ocultan las columnas
        # secundarias menos usadas. El dato sigue disponible en el
        # dataframe subyacente, solo se oculta de la grilla.
        for column in ("Duracion", "PTE", "KKS/TAG"):
            config[column] = None
    return config


def records_table_column_config(state_column: str = "Estado") -> dict[str, Any]:
    return {
        "Titulo tarea": st.column_config.TextColumn("Titulo tarea", width=330),
        "Cuadrilla": st.column_config.TextColumn("Cuadrilla", width=92),
        "OT": st.column_config.TextColumn("OT", width=92),
        "Ubicacion tecnica": st.column_config.TextColumn("Ubicacion tecnica", width=230),
        "KKS/TAG": st.column_config.TextColumn("KKS/TAG", width=105),
        "Tipo registro": st.column_config.TextColumn("Tipo registro", width=130),
        state_column: st.column_config.TextColumn(state_column, width=116),
        "Comentario": st.column_config.TextColumn("Comentario", width=250),
        "Fecha inicio tarea": st.column_config.TextColumn("Fecha inicio tarea", width=118),
        "Fecha finalizacion tarea": st.column_config.TextColumn("Fecha finalizacion tarea", width=142),
        "Fecha modificacion": st.column_config.TextColumn("Fecha modificacion", width=118),
        "Hora modificacion": st.column_config.TextColumn("Hora modificacion", width=104),
        "Informado por": st.column_config.TextColumn("Informado por", width=150),
        "Empresa": st.column_config.TextColumn("Empresa", width=150),
        "Sector": st.column_config.TextColumn("Sector", width=160),
    }


def delete_records_column_config(state_column: str = "Estado") -> dict[str, Any]:
    config = records_table_column_config(state_column)
    config.update(
        {
            "_advance_id": None,
            "Eliminar": st.column_config.CheckboxColumn("Eliminar", width=76),
        }
    )
    return config


def truncate_display_text(value: Any, limit: int = COMMENT_DISPLAY_LIMIT) -> str:
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return text[: max(limit - 3, 0)].rstrip() + "..."


def hide_task_after_saved(task: dict[str, Any], latest: dict[str, dict[str, Any]]) -> bool:
    return effective_task_status(task, latest) in HIDE_AFTER_SAVE_ACTIONS


def format_date(value: Any) -> str:
    if not value:
        return ""
    try:
        return date.fromisoformat(str(value)[:10]).strftime("%d/%m/%Y")
    except Exception:
        return str(value)


def save_advances(program_id: str, task_ids: list[str], action: str, reason: str, observation: str) -> None:
    profile = st.session_state.profile
    rows = [
        {
            "program_id": program_id,
            "task_id": task_id,
            "action": action,
            "reason": reason,
            "observation": observation,
            "reporter_name": profile["name"],
            "reporter_company": profile.get("company") or "",
            "reporter_sector": profile.get("sector") or "",
        }
        for task_id in task_ids
    ]
    sb_insert("advances", rows)


def save_advance_entries(
    program_id: str,
    entries: list[dict[str, str]],
    reason: str,
    observation: str,
    task_dates: dict[str, dict[str, str]] | None = None,
) -> None:
    profile = st.session_state.profile
    rows = [
        {
            "program_id": program_id,
            "task_id": entry["task_id"],
            "action": entry["action"],
            "reason": reason if entry["action"] in REASON_ACTIONS else "",
            "observation": observation,
            "fecha_inicio_tarea": (task_dates or {}).get(entry["task_id"], {}).get("fecha_inicio_tarea"),
            "fecha_finalizacion_tarea": (task_dates or {}).get(entry["task_id"], {}).get("fecha_finalizacion_tarea"),
            "reporter_name": profile["name"],
            "reporter_company": profile.get("company") or "",
            "reporter_sector": profile.get("sector") or "",
        }
        for entry in entries
    ]
    sb_insert("advances", rows)


def combined_comment(advance: dict[str, Any]) -> str:
    reason = str(advance.get("reason") or "").strip()
    observation = str(advance.get("observation") or "").strip()
    if reason and observation:
        return f"{reason} / {observation}"
    return reason or observation


def hide_kks_for_tasks(tasks: list[dict[str, Any]]) -> bool:
    return bool(tasks) and all(task_area(task) == "DISTRIBUCION" for task in tasks)


def strip_kks_if_distribution(rows: list[dict[str, Any]], tasks: list[dict[str, Any]]) -> list[dict[str, Any]]:
    if not hide_kks_for_tasks(tasks):
        return rows
    return [{key: value for key, value in row.items() if key != "KKS/TAG"} for row in rows]


def order_records_by_area(
    rows: list[dict[str, Any]],
    tasks: list[dict[str, Any]],
    state_column: str,
) -> list[dict[str, Any]]:
    if hide_kks_for_tasks(tasks):
        columns = [
            "_advance_id",
            "Titulo tarea",
            "Cuadrilla",
            "OT",
            "Ubicacion tecnica",
            "Tipo registro",
            state_column,
            "Comentario",
            "Fecha inicio tarea",
            "Fecha finalizacion tarea",
            "Fecha modificacion",
            "Hora modificacion",
            "Informado por",
            "Empresa",
            "Sector",
        ]
    else:
        columns = [
            "_advance_id",
            "Titulo tarea",
            "Tipo registro",
            state_column,
            "OT",
            "Cuadrilla",
            "Ubicacion tecnica",
            "KKS/TAG",
            "Comentario",
            "Fecha inicio tarea",
            "Fecha finalizacion tarea",
            "Fecha modificacion",
            "Hora modificacion",
            "Informado por",
            "Empresa",
            "Sector",
        ]
    return [{column: row.get(column, "") for column in columns if column in row} for row in rows]


def advances_export(tasks: list[dict[str, Any]], advances: list[dict[str, Any]], final_only: bool = False) -> bytes:
    tasks_by_id = {task["id"]: task for task in tasks}
    rows = []
    source = advances
    if final_only:
        seen = set()
        filtered = []
        for advance in sorted(advances, key=lambda item: item.get("created_at", ""), reverse=True):
            if advance["task_id"] in seen or advance.get("action") == "COMENTARIO":
                continue
            seen.add(advance["task_id"])
            filtered.append(advance)
        source = filtered
    for advance in source:
        task = tasks_by_id.get(advance["task_id"], {})
        action = advance.get("action", "")
        original_start = original_start_date(task)
        original_end = original_end_date(task)
        program_change = ""
        display_status = action
        if final_only:
            _, _, program_change = wrike_dates_for_advance(task, advance)
            display_status = status_usuario_for_wrike(action)
        rows.append(
            {
                "OT": ot_text(task.get("nro_ot")),
                "Titulo tarea": task_title(task),
                "Empresa": effective_company(task),
                "Sector": effective_sector(task),
                "Cuadrilla": task.get("cuadrilla", ""),
                "KKS/TAG": task.get("kks_tag", ""),
                "Ubicacion tecnica": task.get("ubicacion_tecnica", ""),
                "Fecha inicio original programa": original_start,
                "Fecha fin original programa": original_end,
                "Duracion": task.get("duracion", ""),
                "Estado": display_status,
                **(
                    {
                        "Modificaciones programa": program_change,
                    }
                    if final_only
                    else {}
                ),
                "Comentario": combined_comment(advance),
                "Fecha inicio tarea": advance_task_start_date(advance),
                "Fecha finalizacion tarea": advance_task_finish_date(advance),
                "Fecha avance": advance_date(advance.get("created_at")),
                "Hora avance": advance_time(advance.get("created_at")),
                "Informado por": advance.get("reporter_name", ""),
                "Empresa informante": advance.get("reporter_company", ""),
            }
        )
    rows = strip_kks_if_distribution(rows, tasks)
    return write_excel(rows, sheet_name="Avances")


def status_usuario_for_wrike(action: str) -> str:
    mapping = {
        "EN CURSO": "EN CURSO",
        "COMPLETADO": "COMPLETADO",
        "REPLANIFICAR": "ON HOLD",
        "EN ESPERA": "ON HOLD",
        "SIN AVANCE": "SIN AVANCE",
    }
    return mapping.get(str(action or "").strip().upper(), str(action or "").strip())


def _date_changed(original: Any, updated: str) -> bool:
    if not updated:
        return False
    original_iso = parse_date(original)
    updated_iso = parse_date(updated)
    if not original_iso:
        return bool(updated_iso)
    return bool(updated_iso and original_iso != updated_iso)


def wrike_dates_for_advance(task: dict[str, Any], advance: dict[str, Any]) -> tuple[str, str, str]:
    action = str(advance.get("action") or "").strip().upper()
    start_date = original_start_date(task)
    end_date = original_end_date(task)
    changed = False

    if action == "EN CURSO":
        start_change_date = advance_task_start_date(advance) or advance_date(advance.get("created_at"))
        if start_change_date:
            changed = _date_changed(task.get("fecha_inicio"), start_change_date)
            start_date = start_change_date
    elif action == "COMPLETADO":
        finish_change_date = advance_task_finish_date(advance) or advance_date(advance.get("created_at"))
        if finish_change_date:
            changed = _date_changed(task.get("fecha_fin"), finish_change_date)
            end_date = finish_change_date
        if not start_date:
            start_date = finish_change_date
            changed = True

    program_change = "Cambia fecha de tarea" if changed else ""
    return start_date, end_date, program_change


def wrike_advances_export(tasks: list[dict[str, Any]], advances: list[dict[str, Any]]) -> bytes:
    tasks_by_id = {task["id"]: task for task in tasks}
    latest = latest_status_by_task(advances)
    rows = []
    for task_id, advance in latest.items():
        task = tasks_by_id.get(task_id)
        if not task:
            continue
        order = ot_text(task.get("nro_ot"))
        if not order:
            continue
        action = str(advance.get("action") or "").strip()
        if not action or action == "COMENTARIO":
            continue
        start_date, end_date, program_change = wrike_dates_for_advance(task, advance)
        wrike_status = status_usuario_for_wrike(action)
        rows.append(
            {
                "Orden": order,
                "Estado": wrike_status,
                "Modificaciones programa": program_change,
                "Comentario": combined_comment(advance),
                "Fecha avance": advance_date(advance.get("created_at")),
                "Hora avance": advance_time(advance.get("created_at")),
                "Fecha inicio tarea": advance_task_start_date(advance),
                "Fecha finalizacion tarea": advance_task_finish_date(advance),
                "Fecha inicio original programa": original_start_date(task),
                "Fecha fin original programa": original_end_date(task),
                "Texto breve": task_title(task),
                "Ubicacion tecnica": task.get("ubicacion_tecnica", ""),
                "Cuadrilla": task.get("cuadrilla", ""),
                "KKS/TAG": task.get("kks_tag", ""),
                "Informado por": advance.get("reporter_name", ""),
                "Empresa informante": advance.get("reporter_company", ""),
            }
        )
    columns = [
        "Orden",
        "Estado",
        "Modificaciones programa",
        "Comentario",
        "Fecha avance",
        "Hora avance",
        "Fecha inicio tarea",
        "Fecha finalizacion tarea",
        "Fecha inicio original programa",
        "Fecha fin original programa",
        "Texto breve",
        "Ubicacion tecnica",
        "Cuadrilla",
        "KKS/TAG",
        "Informado por",
        "Empresa informante",
    ]
    return write_excel(rows, columns=columns, sheet_name="Wrike avances")


def local_today() -> date:
    return datetime.now(LOCAL_TZ).date()


def local_datetime(value: Any) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(LOCAL_TZ)
    except Exception:
        return None


def advance_date(value: Any) -> str:
    if not value:
        return ""
    parsed = local_datetime(value)
    if parsed:
        return parsed.strftime("%d/%m/%Y")
    return str(value)[:10]


def advance_time(value: Any) -> str:
    if not value:
        return ""
    parsed = local_datetime(value)
    if parsed:
        return parsed.strftime("%H:%M")
    return ""


def looks_like_date_column(column: str) -> bool:
    normalized = normalize(column)
    return any(part in normalized for part in ["fecha", "vencimiento", "inicio", "fin", "start", "due"])


def export_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    parsed = parse_date(text)
    if parsed:
        return format_date(parsed)
    return text


def write_excel(rows: list[dict[str, Any]], columns: list[str] | None = None, sheet_name: str = "Datos") -> bytes:
    buffer = io.BytesIO()
    df = pd.DataFrame(rows, columns=columns)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        advance_header_fill = PatternFill("solid", fgColor="9BC2E6")
        advance_cell_fill = PatternFill("solid", fgColor="DDEBF7")
        advance_header_font = Font(bold=True, color="1F4E78")
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "")
            if normalize(header) in ADVANCE_EXPORT_COLUMNS:
                column_cells[0].fill = advance_header_fill
                column_cells[0].font = advance_header_font
                for cell in column_cells[1:]:
                    cell.fill = advance_cell_fill
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, len(header) + 2), 45)
    return buffer.getvalue()


def program_updated_export(tasks: list[dict[str, Any]], advances: list[dict[str, Any]]) -> bytes:
    latest = latest_status_by_task(advances)
    rows: list[dict[str, Any]] = []
    columns: list[str] = []

    for task in tasks:
        raw = task.get("raw") or {}
        if isinstance(raw, dict) and raw:
            row = dict(raw)
        else:
            row = {
                "OT": ot_text(task.get("nro_ot")),
                "Titulo tarea": task_title(task),
                "Trabajo": task.get("tarea", ""),
                "Fecha inicio": format_date(task.get("fecha_inicio")),
                "Fecha fin": format_date(task.get("fecha_fin")),
                "Estado": task.get("estado_programa") or "",
                "Cuadrilla": task.get("cuadrilla") or "",
                "Ubicacion tecnica": task.get("ubicacion_tecnica") or "",
                "KKS/TAG": task.get("kks_tag") or "",
            }

        advance = latest.get(task["id"])
        if advance:
            state_col = raw_column_name(row, STATUS_COLUMNS) or "Estado"
            row[state_col] = advance.get("action", "")

            changed_on = advance_operational_date(advance)
            if advance.get("action") == "EN CURSO" and changed_on:
                start_col = raw_column_name(row, START_DATE_COLUMNS) or "Fecha inicio"
                row[start_col] = changed_on
            elif advance.get("action") == "COMPLETADO" and changed_on:
                end_col = raw_column_name(row, END_DATE_COLUMNS) or "Fecha fin"
                row[end_col] = changed_on
                start_col = raw_column_name(row, START_DATE_COLUMNS) or "Fecha inicio"
                if not str(row.get(start_col) or "").strip():
                    row[start_col] = changed_on

        if "OT" in row:
            row["OT"] = ot_text(row.get("OT"))
        else:
            nro_col = raw_column_name(row, OT_COLUMNS)
            if nro_col:
                row[nro_col] = ot_text(row.get(nro_col))

        for column in list(row.keys()):
            if looks_like_date_column(column):
                row[column] = export_date(row.get(column))

        for column in row.keys():
            if column not in columns:
                columns.append(column)
        rows.append(row)

    return write_excel(rows, columns=columns, sheet_name="Programa actualizado")


def render_records_section(
    tasks: list[dict[str, Any]],
    advances: list[dict[str, Any]],
    filtered_base: list[dict[str, Any]],
    filtered_advances: list[dict[str, Any]],
    records_key_suffix: str = "",
) -> None:
    key_suffix = f"_{records_key_suffix}" if records_key_suffix else ""
    title_col, view_col = st.columns([2, 1])
    title_col.subheader("Avances / registros")
    records_view = view_col.selectbox(
        "Vista",
        ["Log completo", "Estado final por OT"],
        key=f"records_view{key_suffix}",
    )
    tasks_by_id = {task["id"]: task for task in tasks}
    source_advances = filtered_advances
    if records_view == "Estado final por OT":
        source_advances = list(latest_status_by_task(filtered_advances).values())
    state_column = "Estado final" if records_view == "Estado final por OT" else "Estado"
    log_rows = [
        {
            "_advance_id": advance.get("id", ""),
            "OT": ot_text(task.get("nro_ot")),
            "Titulo tarea": task_title(task),
            "Fecha modificacion": advance_date(advance.get("created_at")),
            "Hora modificacion": advance_time(advance.get("created_at")),
            "Tipo registro": advance_record_type(advance),
            state_column: advance.get("action"),
            "Comentario": combined_comment(advance),
            "Fecha inicio tarea": advance_task_start_date(advance),
            "Fecha finalizacion tarea": advance_task_finish_date(advance),
            "Informado por": advance.get("reporter_name"),
            "Empresa": effective_company(task),
            "Sector": effective_sector(task),
            "Cuadrilla": task.get("cuadrilla", ""),
            "Ubicacion tecnica": task.get("ubicacion_tecnica", ""),
            "KKS/TAG": task.get("kks_tag", ""),
        }
        for advance in source_advances
        for task in [tasks_by_id.get(advance["task_id"], {})]
    ]
    display_log_rows = order_records_by_area(
        strip_kks_if_distribution(log_rows, filtered_base),
        filtered_base,
        state_column,
    )
    log = pd.DataFrame([{key: value for key, value in row.items() if key != "_advance_id"} for row in display_log_rows])
    display_log = log.copy()
    if "Comentario" in display_log.columns:
        display_log["Comentario"] = display_log["Comentario"].map(truncate_display_text)
    if display_log.empty:
        st.dataframe(display_log, hide_index=True, use_container_width=False, column_config=records_table_column_config(state_column))
    else:
        st.dataframe(
            display_log.style.apply(record_status_style, axis=1),
            hide_index=True,
            use_container_width=False,
            column_config=records_table_column_config(state_column),
        )

    if st.session_state.role == "admin" and advances:
        with st.expander("Administracion de registros", expanded=False):
            st.caption(
                "La limpieza trabaja sobre los registros seleccionados. Puede conservar el ultimo cambio "
                "de estado y/o el ultimo comentario independiente de cada tarea segun la opcion elegida."
            )
            if display_log_rows:
                # Si cambia el conjunto de registros visibles, se destilda el
                # check maestro y se limpia la seleccion del editor de borrado.
                visible_advance_ids = tuple(str(row.get("_advance_id", "")) for row in display_log_rows)
                snapshot_key = f"advances_visible_snapshot{key_suffix}_{records_view}"
                delete_editor_key = f"delete_advances_editor_{records_view}{key_suffix}"
                if st.session_state.get(snapshot_key) != visible_advance_ids:
                    st.session_state.pop(f"select_all_visible_advances{key_suffix}", None)
                    st.session_state.pop(delete_editor_key, None)
                    st.session_state[snapshot_key] = visible_advance_ids
                select_all_visible = st.checkbox(
                    "Seleccionar todo visible",
                    key=f"select_all_visible_advances{key_suffix}",
                )
                delete_df = pd.DataFrame([{**{"Eliminar": False}, **row} for row in display_log_rows])
                if select_all_visible:
                    delete_df["Eliminar"] = True
                edited_delete = st.data_editor(
                    delete_df,
                    hide_index=True,
                    use_container_width=False,
                    disabled=[column for column in delete_df.columns if column != "Eliminar"],
                    column_config=delete_records_column_config(state_column),
                    key=f"delete_advances_editor_{records_view}{key_suffix}",
                )
                selected_delete_ids = (
                    edited_delete.loc[edited_delete["Eliminar"] == True, "_advance_id"].tolist()
                    if not edited_delete.empty
                    else []
                )
            else:
                st.info("No hay registros visibles con los filtros actuales.")
                selected_delete_ids = []

            def run_selected_cleanup(label: str, delete_status: bool, delete_comments: bool) -> None:
                delete_ids, keep_ids = delete_candidates_preserving_latest_records(
                    advances,
                    selected_delete_ids,
                    delete_status=delete_status,
                    delete_comments=delete_comments,
                )
                delete_advances(delete_ids)
                st.session_state.pop(f"delete_advances_editor_{records_view}{key_suffix}", None)
                st.session_state.pop(f"select_all_visible_advances{key_suffix}", None)
                if delete_ids:
                    st.success(
                        f"{label}: {len(delete_ids)} registro(s) eliminado(s). "
                        f"Se conservaron {len(keep_ids)} ultimo(s) registro(s) protegido(s)."
                    )
                else:
                    st.info(f"{label}: no habia registros antiguos para eliminar con esta seleccion.")
                st.rerun()

            d1, d2, d3 = st.columns(3)
            if d1.button(
                "Borrar estados y comentarios antiguos",
                disabled=not selected_delete_ids,
                key=f"clean_status_comments{key_suffix}",
            ):
                run_selected_cleanup("Estados y comentarios", delete_status=True, delete_comments=True)
            if d2.button(
                "Borrar cambios de estado antiguos",
                disabled=not selected_delete_ids,
                key=f"clean_status_only{key_suffix}",
            ):
                run_selected_cleanup("Cambios de estado", delete_status=True, delete_comments=False)
            if d3.button(
                "Borrar comentarios antiguos",
                disabled=not selected_delete_ids,
                key=f"clean_comments_only{key_suffix}",
            ):
                run_selected_cleanup("Comentarios", delete_status=False, delete_comments=True)
            confirm_all = st.checkbox(
                "Confirmo resetear avances del programa activo al estado original del Excel",
                key=f"confirm_reset_advances{key_suffix}",
            )
            if st.button("Resetear al Excel original", disabled=not confirm_all, key=f"reset_advances{key_suffix}"):
                delete_advances([advance["id"] for advance in advances])
                st.session_state.pop(f"delete_advances_editor_{records_view}{key_suffix}", None)
                st.session_state.pop(f"select_all_visible_advances{key_suffix}", None)
                st.success("Registros del programa eliminados.")
                st.rerun()

    e1, e2 = st.columns(2)
    e1.download_button(
        "Exportar log Excel",
        data=advances_export(filtered_base, filtered_advances, final_only=False),
        file_name=f"avances_{local_today().isoformat()}.xlsx",
        key=f"export_log{key_suffix}",
    )
    e2.download_button(
        "Exportar estado final Excel",
        data=advances_export(filtered_base, filtered_advances, final_only=True),
        file_name=f"estado_final_{local_today().isoformat()}.xlsx",
        key=f"export_final{key_suffix}",
    )


def format_datetime(value: Any) -> str:
    if not value:
        return ""
    parsed = local_datetime(value)
    if parsed:
        return parsed.strftime("%d/%m/%Y %H:%M")
    return str(value)


STATE_LABELS = {
    "EN CURSO": "En curso",
    "EN ESPERA": "En espera",
    "COMPLETADO": "Completado",
    "REPLANIFICAR": "A replanificar",
    "SIN AVANCE": "Sin avance",
}


def _is_overdue(task: dict[str, Any], status: str) -> bool:
    if status in {"COMPLETADO"}:
        return False
    end = str(task.get("fecha_fin") or "").strip()
    if not end:
        return False
    try:
        return end[:10] < local_today().isoformat()
    except Exception:
        return False


def render_dashboard(tasks: list[dict[str, Any]], latest: dict[str, dict[str, Any]]) -> None:
    if not tasks:
        return

    statuses = {task["id"]: effective_task_status(task, latest) for task in tasks}
    total = len(tasks)
    counts = {action: 0 for action in STATE_ACTIONS}
    for status in statuses.values():
        counts[status] = counts.get(status, 0) + 1
    completed = counts.get("COMPLETADO", 0)
    overdue = sum(1 for task in tasks if _is_overdue(task, statuses[task["id"]]))
    pct_completed = round((completed / total) * 100) if total else 0

    with st.expander("Resumen del programa", expanded=False):
        m1, m2, m3, m4, m5 = st.columns(5)
        m1.metric("Tareas", total)
        m2.metric("Completadas", f"{completed} ({pct_completed}%)")
        m3.metric("En curso", counts.get("EN CURSO", 0))
        m4.metric("En espera / replanificar", counts.get("EN ESPERA", 0) + counts.get("REPLANIFICAR", 0))
        m5.metric("Vencidas sin completar", overdue)
        if overdue:
            m5.caption(":red[Requieren atencion]")

        chart_col, table_col = st.columns([1.1, 1.4])

        state_df = pd.DataFrame(
            {
                "Estado": [STATE_LABELS.get(action, action) for action in STATE_ACTIONS],
                "Tareas": [counts.get(action, 0) for action in STATE_ACTIONS],
            }
        ).set_index("Estado")
        chart_col.caption("Tareas por estado")
        chart_col.bar_chart(state_df, use_container_width=True)

        breakdown_rows: dict[tuple[str, str], dict[str, int]] = {}
        for task in tasks:
            key = (effective_company(task) or "Sin empresa", effective_sector(task) or "Sin sector")
            entry = breakdown_rows.setdefault(key, {"Tareas": 0, "Completadas": 0})
            entry["Tareas"] += 1
            if statuses[task["id"]] == "COMPLETADO":
                entry["Completadas"] += 1

        table_col.caption("Avance por empresa / sector")
        breakdown_df = pd.DataFrame(
            [
                {
                    "Empresa": company,
                    "Sector": sector,
                    "Tareas": data["Tareas"],
                    "Completadas": data["Completadas"],
                    "% Completado": round((data["Completadas"] / data["Tareas"]) * 100) if data["Tareas"] else 0,
                }
                for (company, sector), data in sorted(breakdown_rows.items())
            ]
        )
        table_col.dataframe(breakdown_df, hide_index=True, use_container_width=True)


def main() -> None:
    require_session()

    if st.session_state.pop("clear_comment_text_next", False):
        st.session_state.pop("common_comment_text", None)
        st.session_state.pop("common_reason_select", None)
        st.session_state.pop("state_select_for_selected", None)
        st.session_state["comment_dirty"] = False
    app_header()
    logout_controls()
    admin_panel()

    profile = st.session_state.profile
    area = canonical_area(profile.get("area") or "GENERACION")
    programs = [program for program in list_programs(active_only=True) if program_area(program) == area]
    if not programs:
        st.info(f"No hay programas activos para {area}. Ingresar como administrador y cargar un Excel.")
        return

    program_col, _program_spacer = st.columns([2.1, 2.9])
    program = program_col.selectbox(
        "Programa",
        programs,
        format_func=lambda item: item["name"],
        key="program_filter",
    )
    tasks = load_tasks(program["id"])
    advances = load_advances(program["id"])
    latest = latest_status_by_task(advances)

    company = scope_value(profile.get("company"))
    sector = scope_value(profile.get("sector"))
    scoped_tasks = apply_filters(tasks, company, sector, "", None, None, "")
    if st.session_state.role == "admin" and profile.get("admin_view") == "Avances / registros":
        scoped_task_ids = {task["id"] for task in scoped_tasks}
        scoped_advances = [advance for advance in advances if advance.get("task_id") in scoped_task_ids]
        st.caption(
            f"Vista de administrador: {area} - {profile.get('company') or 'Todas'} / "
            f"{profile.get('sector') or 'Todos'}. Los registros se mantienen separados por area."
        )
        render_records_section(tasks, advances, scoped_tasks, scoped_advances, records_key_suffix="admin_direct")
        return

    inner_sector_options = [""] + sorted(
        {effective_sector(task) for task in scoped_tasks if effective_sector(task)},
        key=lambda item: normalize(item),
    )
    f1, f2, f3, f4, f5 = st.columns([0.95, 1.25, 0.62, 0.62, 1.95])
    inner_sector = f1.selectbox(
        "Sector",
        inner_sector_options,
        format_func=lambda item: option_label(item, "Todos"),
        key="sector_filter",
    )
    tasks_for_crews = apply_filters(scoped_tasks, "", inner_sector, "", None, None, "")
    crew_options = sorted({task.get("cuadrilla") or "" for task in tasks_for_crews if task.get("cuadrilla")})
    if "crew_filter" in st.session_state and not isinstance(st.session_state.get("crew_filter"), list):
        st.session_state.pop("crew_filter", None)
    crew = f2.multiselect("Cuadrilla", crew_options, placeholder="Todas", key="crew_filter")
    start = f3.date_input("Fecha inicio", value=None, format="DD/MM/YYYY", key="start_filter")
    end = f4.date_input("Fecha fin", value=None, format="DD/MM/YYYY", key="end_filter")
    active_search_terms = st.session_state.setdefault("search_terms_filter", [])
    search_field_col, search_value_col = f5.columns([0.9, 1.25])
    search_field_col.selectbox("Buscar por", list(SEARCH_FIELD_OPTIONS.keys()), key="search_field_input")
    search_value_col.text_input(
        "Buscar",
        placeholder="Escribir y presionar Enter",
        key="search_value_input",
        on_change=add_search_filter_from_state,
    )
    active_search_terms = st.session_state.get("search_terms_filter", [])
    if active_search_terms:
        tag_cols = st.columns(4)
        for index, label in enumerate(active_search_terms):
            if tag_cols[index % 4].button(f"x {label}", key=f"remove_search_filter_{index}_{normalize(label)}"):
                remove_search_filter(label)
                st.rerun()
    show_all_tasks = bool(st.session_state.get("show_all_tasks_filter", False))

    filters_now = current_filter_state(program["id"])

    filters_before = st.session_state.get("last_filter_state")
    if filters_before is None:
        st.session_state.last_filter_state = filters_now
    elif filters_now != filters_before:
        if has_pending_work():
            st.session_state.filter_change_guard = {
                "previous": filters_before,
                "current": filters_now,
            }
        else:
            st.session_state.previous_filter_state = filters_before
            st.session_state.last_filter_state = filters_now
            # Cambio el conjunto visible: se limpia la seleccion y se destilda
            # el check maestro (no queda ninguna tarea seleccionada).
            clear_task_selection()

    base_sector = inner_sector or sector
    filtered_base = apply_filters(tasks, company, base_sector, crew, start, end, active_search_terms)
    loaded_for_scope = apply_filters(tasks, company, base_sector, "", None, None, "")

    render_dashboard(filtered_base, latest)
    render_crew_status_section(area, crew_options, company, base_sector, crew_context_lookup(tasks))
    if show_all_tasks:
        filtered = filtered_base
    else:
        filtered = [task for task in filtered_base if not hide_task_after_saved(task, latest)]
    hidden_count = len(filtered_base) - len(filtered)
    caption = f"{len(filtered)} tarea(s) visibles de {len(loaded_for_scope)} cargadas para empresa/sector."
    if hidden_count:
        caption += f" {hidden_count} completada(s) o a replanificar ocultas."
    visible_task_ids = {task["id"] for task in filtered}
    filtered_base_task_ids = {task["id"] for task in filtered_base}
    # Los avances/registros siguen los MISMOS filtros que el tablero (empresa,
    # sector, cuadrilla, fechas, busqueda). Se incluyen las tareas completadas o
    # a replanificar que se ocultan del tablero, para no perder sus registros.
    filtered_advances = [advance for advance in advances if advance.get("task_id") in filtered_base_task_ids]
    scoped_task_ids = {task["id"] for task in scoped_tasks}
    scoped_advances = [advance for advance in advances if advance.get("task_id") in scoped_task_ids]

    pending_changes = st.session_state.setdefault("pending_state_changes", {})
    selected_task_state = set(st.session_state.setdefault("selected_task_ids", []))

    # Manejo del check maestro "Seleccionar tareas visibles". Solo actua en la
    # TRANSICION (cuando se tilda o se destilda), no en cada rerun, para que las
    # tareas que el usuario destilde individualmente queden respetadas y no se
    # vuelvan a marcar solas.
    master_now = bool(st.session_state.get("select_all_visible_tasks_filter", False))
    master_prev = bool(st.session_state.get("select_all_visible_tasks_prev", False))
    if master_now and not master_prev:
        # Recien tildado: seleccionar todas las visibles (una sola vez). Se
        # descartan ediciones de "Seleccionar" previas del editor; los cambios
        # de estado ya viven en pending_changes y se reaplican mas abajo.
        selected_task_state.update(visible_task_ids)
        st.session_state.pop("task_editor", None)
    elif master_prev and not master_now:
        # Recien destildado: limpiar toda la seleccion.
        selected_task_state.clear()
        st.session_state.pop("task_editor", None)
    elif master_now and visible_task_ids and not visible_task_ids.issubset(selected_task_state):
        # Sigue tildado pero el usuario destildo alguna: destildar el maestro
        # (ya no estan todas las visibles) sin tocar la seleccion restante.
        st.session_state["select_all_visible_tasks_filter"] = False
        master_now = False
    st.session_state.select_all_visible_tasks_prev = master_now
    st.session_state.selected_task_ids = sorted(selected_task_state)

    df = task_dataframe(filtered, latest)
    for index, row in df.iterrows():
        task_id = str(row["_task_id"])
        if task_id in selected_task_state:
            df.at[index, "Seleccionar"] = True
        if task_id in pending_changes:
            df.at[index, "Estado"] = pending_changes[task_id]
            df.at[index, "Seleccionar"] = True
            if pending_changes[task_id] == "EN CURSO" and not str(df.at[index, "Fecha inicio"] or "").strip():
                df.at[index, "Fecha inicio"] = local_today().strftime("%d/%m/%Y")

    editor_state = st.session_state.get("task_editor", {})
    if isinstance(editor_state, dict):
        for row_index, changes in editor_state.get("edited_rows", {}).items():
            try:
                index = int(row_index)
            except Exception:
                continue
            if 0 <= index < len(df):
                if "Seleccionar" in changes:
                    df.at[index, "Seleccionar"] = bool(changes["Seleccionar"])
                if "Estado" in changes:
                    new_status = str(changes["Estado"] or "")
                    original_status = str(df.at[index, "_estado_original"] or "")
                    df.at[index, "Estado"] = new_status
                    if new_status and new_status != original_status:
                        df.at[index, "Seleccionar"] = True
                        task_id = str(df.at[index, "_task_id"])
                        pending_changes[task_id] = new_status
                        st.session_state.pending_program_id = program["id"]
                        if new_status == "EN CURSO" and not str(df.at[index, "Fecha inicio"] or "").strip():
                            df.at[index, "Fecha inicio"] = local_today().strftime("%d/%m/%Y")

    selected_task_ids = [
        str(row["_task_id"])
        for _, row in df.iterrows()
        if bool(row.get("Seleccionar"))
    ]
    pending_visible_ids = sorted(task_id for task_id in pending_changes if task_id in visible_task_ids)
    pending_all_ids = sorted(pending_changes)

    entries = [{"task_id": task_id, "action": pending_changes[task_id]} for task_id in pending_all_ids]
    selected_actions = [entry["action"] for entry in entries]
    pending_needs_reason = any(item in REASON_ACTIONS for item in selected_actions)

    st.markdown("#### Cambiar estado / comentar")
    start_date_mode = DATE_MODE_PROGRAM
    finish_date_mode = DATE_MODE_PROGRAM
    manual_start_task_date: date | None = None
    manual_finish_task_date: date | None = None
    # Los campos aparecen solo cuando hay tareas seleccionadas (o cambios ya
    # pendientes). El estado y el comentario elegidos aca NO se aplican hasta
    # tocar "Guardar cambios".
    if selected_task_ids or pending_all_ids:
        # Valor previo del desplegable, para decidir si mostrar el motivo antes
        # de instanciar el widget de estado.
        action_prev = str(st.session_state.get("state_select_for_selected", "") or "")
        show_reason = pending_needs_reason or action_prev in REASON_ACTIONS

        if show_reason:
            estado_col, motivo_col, detalle_col = st.columns([1.0, 1.0, 1.8])
        else:
            estado_col, detalle_col = st.columns([1.0, 2.0])
            motivo_col = None

        action = estado_col.selectbox(
            "Estado para seleccionadas",
            [""] + STATE_ACTIONS,
            format_func=lambda item: option_label(item, "Elegir estado"),
            key="state_select_for_selected",
            help="Se aplica a las tareas tildadas al tocar 'Guardar cambios'. Tambien podes cambiar el estado directamente en la columna Estado de cada fila.",
        )
        if motivo_col is not None:
            reason = motivo_col.selectbox(
                "Motivo (obligatorio)",
                REASONS,
                format_func=lambda item: option_label(item, "Elegir motivo"),
                key="common_reason_select",
            )
        else:
            reason = ""

        # El motivo/detalle obligatorio depende de lo que se va a guardar:
        # pendientes con EN ESPERA/REPLANIFICAR, o el estado elegido ahora.
        will_need_reason = pending_needs_reason or action in REASON_ACTIONS
        detail_required = will_need_reason and reason == "Otros"
        detail_label = "Detalle (obligatorio)" if detail_required else "Detalle (opcional)"
        observation = detalle_col.text_input(
            detail_label,
            placeholder="Detalle libre. Obligatorio si el motivo es 'Otros'. Tambien sirve como comentario de las tareas seleccionadas.",
            key="common_comment_text",
            on_change=mark_comment_dirty,
        )

        planned_actions = set(selected_actions)
        if action:
            planned_actions.add(action)
        date_cols: list[Any] = []
        if "EN CURSO" in planned_actions and "COMPLETADO" in planned_actions:
            date_cols = list(st.columns(2))
        elif "EN CURSO" in planned_actions or "COMPLETADO" in planned_actions:
            date_cols = [st.container()]

        if "EN CURSO" in planned_actions:
            with date_cols.pop(0):
                start_date_mode = st.selectbox(
                    "Fecha inicio tarea",
                    TASK_DATE_MODES,
                    index=TASK_DATE_MODES.index(DATE_MODE_PROGRAM),
                    key="start_task_date_mode",
                    help="Fecha real de inicio para las tareas que se informan EN CURSO.",
                )
                if start_date_mode == DATE_MODE_MANUAL:
                    manual_start_task_date = st.date_input(
                        "Ingresar fecha inicio tarea",
                        value=local_today(),
                        format="DD/MM/YYYY",
                        key="manual_start_task_date",
                    )
        if "COMPLETADO" in planned_actions:
            with date_cols.pop(0) if date_cols else st.container():
                finish_date_mode = st.selectbox(
                    "Fecha finalizacion tarea",
                    TASK_DATE_MODES,
                    index=TASK_DATE_MODES.index(DATE_MODE_PROGRAM),
                    key="finish_task_date_mode",
                    help="Fecha real de finalizacion para las tareas que se informan COMPLETADO.",
                )
                if finish_date_mode == DATE_MODE_MANUAL:
                    manual_finish_task_date = st.date_input(
                        "Ingresar fecha finalizacion tarea",
                        value=local_today(),
                        format="DD/MM/YYYY",
                        key="manual_finish_task_date",
                    )
    else:
        action = ""
        reason = ""
        observation = ""
        st.caption("Selecciona una o mas tareas para cambiar su estado o dejar un comentario.")

    def save_current_pending_work() -> bool:
        save_program_id = st.session_state.get("pending_program_id") or program["id"]
        # Aplicar el estado elegido en el desplegable a las tareas seleccionadas.
        if action and selected_task_ids:
            for task_id in selected_task_ids:
                pending_changes[task_id] = action
            save_program_id = program["id"]
        current_pending_ids = sorted(pending_changes)
        current_entries = [{"task_id": tid, "action": pending_changes[tid]} for tid in current_pending_ids]
        current_actions = [item["action"] for item in current_entries]
        needs_reason = any(item in REASON_ACTIONS for item in current_actions)
        if current_pending_ids:
            if needs_reason and not reason:
                st.warning("Elegi un motivo para EN ESPERA o REPLANIFICAR.")
                return False
            if needs_reason and reason == "Otros" and not observation.strip():
                st.warning("Escribi el detalle cuando el motivo es 'Otros'.")
                return False
            tasks_by_id_for_dates = {str(task["id"]): task for task in tasks}
            task_dates: dict[str, dict[str, str]] = {}
            for entry in current_entries:
                task_id = str(entry["task_id"])
                task = tasks_by_id_for_dates.get(task_id, {})
                action_for_date = str(entry["action"] or "").strip().upper()
                if action_for_date == "EN CURSO":
                    start_task_date = task_date_by_mode(
                        task,
                        start_date_mode,
                        manual_start_task_date,
                        "fecha_inicio",
                    )
                    if not start_task_date:
                        st.warning(
                            "Hay una tarea EN CURSO sin fecha de inicio en el programa. "
                            "Elegí 'Fecha de hoy' o 'Fecha manual'."
                        )
                        return False
                    task_dates[task_id] = {
                        "fecha_inicio_tarea": start_task_date
                    }
                elif action_for_date == "COMPLETADO":
                    finish_task_date = task_date_by_mode(
                        task,
                        finish_date_mode,
                        manual_finish_task_date,
                        "fecha_fin",
                    )
                    if not finish_task_date:
                        st.warning(
                            "Hay una tarea COMPLETADA sin fecha de fin en el programa. "
                            "Elegí 'Fecha de hoy' o 'Fecha manual'."
                        )
                        return False
                    task_dates[task_id] = {
                        "fecha_finalizacion_tarea": finish_task_date
                    }
            save_advance_entries(save_program_id, current_entries, reason, observation.strip(), task_dates)
            clear_pending_work()
            return True
        if observation.strip() and selected_task_ids:
            comment_entries = [{"task_id": task_id, "action": "COMENTARIO"} for task_id in selected_task_ids]
            save_advance_entries(program["id"], comment_entries, "", observation.strip())
            clear_pending_work()
            return True
        st.warning("No hay cambios ni comentarios para guardar.")
        return False

    # Hay algo para guardar si hay pendientes, un estado elegido para las
    # seleccionadas, o un comentario para las seleccionadas.
    has_savable_changes = (
        bool(pending_all_ids)
        or bool(action and selected_task_ids)
        or bool(observation.strip() and selected_task_ids)
    )

    if st.session_state.pop("request_refresh", False):
        if pending_all_ids:
            st.session_state.confirm_refresh = True
        else:
            clear_task_selection()
            st.rerun()

    if st.session_state.get("confirm_refresh"):
        st.warning("Hay avances pendientes sin guardar. Elegi como continuar antes de actualizar datos.")
        r1, r2, r3 = st.columns(3)
        if r1.button("Guardar avances y actualizar", type="primary"):
            if save_current_pending_work():
                st.rerun()
        if r2.button("Actualizar sin guardar"):
            clear_pending_work()
            st.rerun()
        if r3.button("Cancelar"):
            st.session_state.pop("confirm_refresh", None)
            st.rerun()

    filter_guard = st.session_state.get("filter_change_guard")
    if filter_guard:
        st.warning("Cambiaste un filtro y hay avances o comentarios pendientes sin guardar.")
        f1, f2, f3 = st.columns(3)

        if f1.button("Guardar y aplicar filtro", type="primary"):
            previous = filter_guard.get("previous")
            current = filter_guard.get("current")
            if save_current_pending_work():
                st.session_state.previous_filter_state = previous
                st.session_state.last_filter_state = current
                st.session_state.pop("filter_change_guard", None)
                st.rerun()

        def apply_filter_without_saving() -> None:
            guard = st.session_state.get("filter_change_guard") or {}
            previous = guard.get("previous")
            current = guard.get("current")
            # El usuario confirma el filtro nuevo y descarta todo avance/comentario
            # pendiente. Como esto corre en callback, ocurre antes del nuevo render.
            clear_pending_work()
            st.session_state.previous_filter_state = previous
            st.session_state.last_filter_state = current
            st.session_state.pop("filter_change_guard", None)

        def cancel_filter_change() -> None:
            guard = st.session_state.get("filter_change_guard") or {}
            previous = guard.get("previous")
            if previous:
                # Callback de Streamlit: se ejecuta antes del siguiente render,
                # por lo que es seguro restaurar las keys de los widgets acá.
                restore_filter_state(previous)
                st.session_state.last_filter_state = previous
            st.session_state.pop("filter_change_guard", None)
            st.session_state.pop("previous_filter_state", None)
            clear_task_selection()

        f2.button(
            "Aplicar filtro sin guardar",
            on_click=apply_filter_without_saving,
        )
        f3.button(
            "Cancelar cambio de filtro",
            on_click=cancel_filter_change,
        )

    pending_navigation = st.session_state.get("pending_navigation")
    if pending_navigation:
        label = "volver al inicio" if pending_navigation == "profile" else "cerrar sesion"
        st.warning(f"Hay avances o comentarios pendientes sin guardar antes de {label}.")
        n1, n2, n3 = st.columns(3)
        if n1.button("Guardar y continuar", type="primary"):
            if save_current_pending_work():
                apply_navigation(pending_navigation)
        if n2.button("Continuar sin guardar"):
            clear_pending_work()
            apply_navigation(pending_navigation)
        if n3.button("Cancelar salida"):
            st.session_state.pop("pending_navigation", None)
            st.rerun()

    guard_active = bool(
        st.session_state.get("confirm_refresh")
        or st.session_state.get("filter_change_guard")
        or st.session_state.get("pending_navigation")
    )

    save_col, discard_col, pending_col, select_all_col, show_col = st.columns([1.1, 1.15, 1.05, 1.25, 1.2])
    if not guard_active:
        if save_col.button("Guardar cambios", type="primary", disabled=not has_savable_changes, use_container_width=True):
            # Contar lo que se va a guardar antes de aplicar/limpiar.
            affected = set(pending_all_ids)
            if action and selected_task_ids:
                affected.update(selected_task_ids)
            comment_only = not affected and bool(observation.strip()) and bool(selected_task_ids)
            if save_current_pending_work():
                if affected:
                    st.success(f"{len(affected)} tarea(s) actualizada(s).")
                elif comment_only:
                    st.success(f"{len(selected_task_ids)} comentario(s) guardado(s).")
                else:
                    st.success("Cambios guardados.")
                st.rerun()

        # "No guardar cambios": descarta los cambios de estado y/o comentarios
        # pendientes y vuelve al estado previo. Se habilita cuando hay algo que
        # descartar (cambio de estado pendiente, estado elegido o comentario).
        if discard_col.button("No guardar cambios", disabled=not (has_pending_work() or has_savable_changes), use_container_width=True):
            clear_pending_work()
            st.rerun()

        if pending_all_ids:
            pending_col.caption(f"{len(pending_all_ids)} avance(s) pendiente(s) de guardar.")
    else:
        # Hay un aviso activo arriba: resolverlo con sus propios botones.
        save_col.caption("Resolve el aviso de arriba para continuar.")
    select_all_col.checkbox(
        "Seleccionar tareas visibles",
        key="select_all_visible_tasks_filter",
    )
    show_col.checkbox("Mostrar todas las tareas", value=show_all_tasks, key="show_all_tasks_filter")
    compact_view = st.checkbox(
        "Vista compacta (recomendada en celular)",
        key="compact_view_filter",
        help="Muestra Estado, Titulo, Fecha inicio, Cuadrilla, OT y Ubicacion tecnica, pensado para pantallas chicas. Oculta Duracion, PTE y KKS/TAG (se pueden ver activando la vista completa).",
    )
    st.caption(caption)

    visible_df = df.drop(columns=["_task_id", "_estado_original"], errors="ignore")
    edited = st.data_editor(
        visible_df.style.apply(task_status_style, axis=1),
        hide_index=True,
        use_container_width=False,
        disabled=[column for column in visible_df.columns if column not in {"Seleccionar", "Estado"}],
        column_config=task_table_column_config(compact=compact_view),
        key="task_editor",
    )
    if not edited.empty:
        previous_visible_selection = selected_task_state.intersection(visible_task_ids)
        selected_task_state.difference_update(visible_task_ids)
        current_visible_selection = {
            str(df.at[index, "_task_id"])
            for index in edited.index[edited["Seleccionar"] == True].tolist()
            if index in df.index
        }
        selected_task_state.update(current_visible_selection)
        for index, row in edited.iterrows():
            if index not in df.index:
                continue
            task_id = str(df.at[index, "_task_id"])
            status = str(row["Estado"] or "")
            original_status = str(df.at[index, "_estado_original"] or "")
            if status and status != original_status:
                pending_changes[task_id] = status
                st.session_state.pending_program_id = program["id"]
            elif task_id in pending_changes:
                pending_changes.pop(task_id, None)
        st.session_state.selected_task_ids = sorted(selected_task_state)
        if current_visible_selection != previous_visible_selection:
            st.rerun()

    render_records_section(tasks, advances, filtered_base, filtered_advances)


if __name__ == "__main__":
    main()
